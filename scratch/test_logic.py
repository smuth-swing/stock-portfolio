
import openpyxl
from openpyxl.styles import PatternFill

def test_portfolio_sync(file_path, stock_name, new_total_amount, trade_type):
    print(f"--- Testing {trade_type} for {stock_name} with new total {new_total_amount} ---")
    wb = openpyxl.load_workbook(file_path)
    if '포트폴리오 맵' not in wb.sheetnames:
        print("Error: Sheet '포트폴리오 맵' not found")
        return
    
    ws_map = wb['포트폴리오 맵']
    target_row = None
    trade_stock_clean = stock_name.replace(" ", "")
    
    for r in range(1, ws_map.max_row + 1):
        cell_val = str(ws_map.cell(row=r, column=4).value or "").strip()
        cell_val_clean = cell_val.replace(" ", "")
        if cell_val_clean and (trade_stock_clean in cell_val_clean or cell_val_clean in trade_stock_clean):
            target_row = r
            break
            
    if not target_row:
        print(f"Error: Stock {stock_name} not found in Portfolio Map")
        return

    # Count current ones
    current_ones = 0
    for c in range(5, 100):
        cell = ws_map.cell(row=target_row, column=c)
        if cell.value is not None and str(cell.value).strip() != "":
            current_ones += 1
            
    print(f"Current ones in map: {current_ones}")
    
    target_ones = int(new_total_amount // 100)
    print(f"Target ones (total): {target_ones}")
    
    diff = target_ones - current_ones
    print(f"Difference to apply: {diff}")
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    no_fill = PatternFill(fill_type=None)
    
    if diff > 0:
        added = 0
        for c in range(5, 100):
            cell = ws_map.cell(row=target_row, column=c)
            if cell.value is None or str(cell.value).strip() == "":
                cell.value = 1
                cell.fill = yellow_fill
                added += 1
                if added >= diff: break
        print(f"Added {added} ones")
    elif diff < 0:
        removed = 0
        for c in range(100, 4, -1):
            cell = ws_map.cell(row=target_row, column=c)
            if cell.value is not None and str(cell.value).strip() != "":
                cell.value = None
                cell.fill = no_fill
                removed += 1
                if removed >= abs(diff): break
        print(f"Removed {removed} ones")
    else:
        print("No change needed")
        
    wb.save("test_output.xlsx")
    print("Saved to test_output.xlsx")

# Mock Excel for testing
def create_mock_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '포트폴리오 맵'
    ws.cell(row=1, column=4, value="종목")
    ws.cell(row=2, column=4, value="리가켐바이오")
    # Add 2 existing ones (representing 200)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.cell(row=2, column=5, value=1).fill = yellow_fill
    ws.cell(row=2, column=6, value=1).fill = yellow_fill
    wb.save("test_portfolio.xlsx")

if __name__ == "__main__":
    create_mock_excel()
    # Scenario: Current 200, buy 1,000,000 KRW -> new total 300
    test_portfolio_sync("test_portfolio.xlsx", "리가켐바이오", 300, "매수")
    
    # Check output
    wb = openpyxl.load_workbook("test_output.xlsx")
    ws = wb['포트폴리오 맵']
    final_count = 0
    for c in range(5, 100):
        if ws.cell(row=2, column=c).value == 1:
            final_count += 1
    print(f"Final count in output: {final_count}")
    assert final_count == 3
    print("Test Passed!")
