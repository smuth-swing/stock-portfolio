"""한세실업, AJ네트웍스의 Git 이력별 내용 변화 상세 추적"""
import subprocess
import json
import sys
sys.stdout.reconfigure(encoding='utf-8')

CWD = r'c:\Users\zerod\.antigravity\주식 포트폴리오 관리'

# investigation.json이 변경된 모든 커밋
result = subprocess.run(
    ['git', 'log', '--oneline', '--', 'StockPortfolioApp/public/data/investigation.json'],
    capture_output=True, text=True, encoding='utf-8', cwd=CWD
)
all_commits = []
for line in result.stdout.strip().split('\n'):
    if line.strip():
        parts = line.strip().split(' ', 1)
        all_commits.append((parts[0], parts[1] if len(parts) > 1 else ''))

TARGET_NAMES = ['한세실업', 'AJ네트웍스', 'aj네트웍스']

print("=" * 80)
print("  한세실업 & AJ네트웍스 — Git 이력 전수 추적")
print("=" * 80)

for commit, msg in all_commits:
    try:
        r = subprocess.run(
            ['git', 'show', f'{commit}:StockPortfolioApp/public/data/investigation.json'],
            capture_output=True, text=True, encoding='utf-8', cwd=CWD
        )
        if r.returncode != 0:
            continue
        
        data = json.loads(r.stdout)
        cols = data['columns']
        
        for row in data['data']:
            # 종목명 찾기 - 여러 가능한 컬럼에서
            name = ''
            for col in cols:
                val = str(row.get(col, '')).strip()
                for target in TARGET_NAMES:
                    if target in val:
                        name = target
                        break
                if name:
                    break
            
            if not name:
                continue
            
            # 해당 종목의 전체 내용 출력
            print(f"\n{'─' * 80}")
            print(f"📌 커밋: {commit} | {msg}")
            print(f"   종목: {name}")
            for col in cols:
                val = str(row.get(col, '')).strip()
                if val:
                    # 긴 내용은 줄바꿈 표시
                    val_display = val.replace('\\n', '\n         ')
                    print(f"   {col}: {val_display}")
    except Exception as e:
        pass

# 현재 엑셀 파일에서도 확인
print(f"\n{'═' * 80}")
print("  현재 OneDrive 엑셀 원본에서의 상태")
print(f"{'═' * 80}")

import io
import pandas as pd
import openpyxl
from openpyxl.cell.rich_text import CellRichText

excel_path = r'C:\Users\zerod\OneDrive\주식 체크 리스트_20220328.xlsx'
with open(excel_path, 'rb') as f:
    file_data = f.read()

wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
ws = wb['탐구생활']

print(f"시트 행수: {ws.max_row}, 열수: {ws.max_column}")
print(f"헤더: ", end="")
for c in range(1, ws.max_column + 1):
    print(f"{ws.cell(1, c).value}", end=" | ")
print()

for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(r, c)
        val = str(cell.value or '').strip()
        if any(t in val for t in TARGET_NAMES):
            print(f"\n행 {r}: {val}")
            print(f"  전체 행 내용:")
            for c2 in range(1, ws.max_column + 1):
                cell2 = ws.cell(r, c2)
                v = cell2.value
                if v is not None:
                    col_name = ws.cell(1, c2).value or f'Col{c2}'
                    print(f"    {col_name}: {v}")
            break

wb.close()
