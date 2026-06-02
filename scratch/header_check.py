import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\zerod\OneDrive\주식 체크 리스트_20220328.xlsx')
ws = wb['탐구생활']
# 헤더 행(row=3) 컬럼명 전체 - 한글 깨짐없이
for c in range(1, 8):
    val = ws.cell(row=3, column=c).value
    print(f'col{c}:', repr(val))

print()
# 데이터 행 몇 개 확인
for r in range(4, 9):
    row_data = [ws.cell(row=r, column=c).value for c in range(1, 8)]
    print(f'row={r}:', row_data)
