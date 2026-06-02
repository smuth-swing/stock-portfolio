import json

with open(r'C:\Users\zerod\.antigravity\주식 포트폴리오 관리\mobile\data\investigation.json', encoding='utf-8') as f:
    data = json.load(f)

# investigation.json의 data 구조 확인 (JSON index vs 엑셀 row 매핑)
print("=== investigation.json data 배열 ===")
for i in range(min(10, len(data['data']))):
    row = data['data'][i]
    unnamed0 = repr(row.get('Unnamed: 0'))
    unnamed1 = repr(row.get('Unnamed: 1', ''))[:40]
    real_idx = row.get('_realIndex')
    print(f"  data[{i}]: _realIndex={real_idx}, Col0={unnamed0}, Col1={unnamed1}")

print()
print("=== 앱 로직: allData.slice(2) → 첫 2행 건너뜀 ===")
allItems = data['data'][2:]
for i, item in enumerate(allItems[:8]):
    computed_realIndex = i + 2
    print(f"  allItems[{i}]: _realIndex={computed_realIndex}, Col0={repr(item.get('Unnamed: 0'))}, Col1={repr(item.get('Unnamed: 1',''))[:40]}")

# 엑셀과 비교
import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\zerod\OneDrive\주식 체크 리스트_20220328.xlsx')
ws = wb['탐구생활']

print()
print("=== 엑셀 실제 행 ===")
for r in range(1, 11):
    col1 = ws.cell(row=r, column=1).value
    col2_raw = ws.cell(row=r, column=2).value
    col2 = repr(col2_raw)[:40] if col2_raw else repr(col2_raw)
    print(f"  excel row={r}: Col1={repr(col1)}, Col2={col2}")

print()
print("=== _realIndex → 엑셀 row 매핑 검증 ===")
for i, item in enumerate(allItems[:8]):
    realIndex = i + 2
    json_col0 = item.get('Unnamed: 0')
    
    # target_row = realIndex + 2 인 경우
    target_row_plus2 = realIndex + 2
    excel_col1_plus2 = ws.cell(row=target_row_plus2, column=1).value if target_row_plus2 <= ws.max_row else 'OUT'
    
    # target_row = realIndex (offset 없이)
    target_row_no_offset = realIndex
    excel_col1_no = ws.cell(row=target_row_no_offset, column=1).value if target_row_no_offset >= 1 else 'OUT'
    
    # target_row = realIndex + 1
    target_row_plus1 = realIndex + 1
    excel_col1_plus1 = ws.cell(row=target_row_plus1, column=1).value if target_row_plus1 <= ws.max_row else 'OUT'
    
    match_str = ""
    if repr(json_col0) == repr(excel_col1_no):
        match_str = "✓ MATCH: offset=0"
    elif repr(json_col0) == repr(excel_col1_plus1):
        match_str = "✓ MATCH: offset=+1"
    elif repr(json_col0) == repr(excel_col1_plus2):
        match_str = "✓ MATCH: offset=+2"
    else:
        match_str = "✗ NO MATCH"
    
    print(f"  realIndex={realIndex}, json_col0={repr(json_col0)}, " +
          f"excel[row={target_row_no_offset}]={repr(excel_col1_no)}, " +
          f"excel[row={target_row_plus1}]={repr(excel_col1_plus1)}, " +
          f"excel[row={target_row_plus2}]={repr(excel_col1_plus2)} → {match_str}")

wb.close()
