import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
import re
import io

def parse_strikethrough_text(text):
    if not isinstance(text, str) or not text:
        return text
    text = re.sub(r'<del>(.*?)</del>', r'~~\1~~', text)
    pattern = r'~~(.*?)~~'
    parts = re.split(pattern, text)
    if len(parts) == 1: return text
    
    rich_text = CellRichText()
    default_font = InlineFont(rFont='맑은 고딕')
    strike_font = InlineFont(rFont='맑은 고딕', strike=True)
    
    for i, part in enumerate(parts):
        if not part: continue
        if i % 2 == 0:
            rich_text.append(TextBlock(default_font, part))
        else:
            rich_text.append(TextBlock(strike_font, part))
    return rich_text

def extract_rich_text(cell):
    if not hasattr(cell, 'value') or cell.value is None:
        return ''
    if isinstance(cell.value, CellRichText):
        result = []
        for part in cell.value:
            text = part.text if hasattr(part, 'text') else str(part)
            if hasattr(part, 'font') and part.font and getattr(part.font, 'strike', False):
                result.append(f'~~{text}~~')
            else:
                result.append(text)
        return ''.join(result)
    if hasattr(cell, 'font') and cell.font and getattr(cell.font, 'strike', False):
        return f'~~{cell.value}~~'
    return str(cell.value)

# 1. Create a workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "탐구생활"
ws['A1'] = "Original Text"

# 2. Simulate update_row
target_row = 1
value = "This is a ~~strikethrough~~ test"
processed_value = parse_strikethrough_text(value)
print("Processed Value Type:", type(processed_value))
ws.cell(row=target_row, column=1).value = processed_value
wb.save('test_integration.xlsx')

# 3. Simulate export_to_json.py reading the file
with open('test_integration.xlsx', 'rb') as f:
    file_data = f.read()

wb_read = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True, rich_text=True)
ws_read = wb_read["탐구생활"]
cell = ws_read['A1']
print("Cell Value Type after read:", type(cell.value))
print("Extracted Text:", extract_rich_text(cell))
