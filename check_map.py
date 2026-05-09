import openpyxl
import glob
import os

try:
    files = glob.glob('*.xlsx')
    if not files:
        print("Excel file not found")
        exit(1)
    
    fname = files[0]
    wb = openpyxl.load_workbook(fname, data_only=True)
    
    if '포트폴리오 맵' not in wb.sheetnames:
        print("Sheet '포트폴리오 맵' not found")
        exit(1)
        
    ws = wb['포트폴리오 맵']
    print(f"--- Checking {fname} ---")
    
    found = False
    for r in range(1, ws.max_row + 1):
        name = str(ws.cell(row=r, column=4).value or "").strip()
        if '리가켐바이오' in name:
            found = True
            marks = 0
            for c in range(5, 101):
                val = ws.cell(row=r, column=c).value
                if val == 1 or str(val) == "1":
                    marks += 1
            print(f"Row {r}: [{name}] | Marks: {marks}")
    
    if not found:
        print("Stock '리가켐바이오' not found in Map")
except Exception as e:
    print(f"Error: {e}")
