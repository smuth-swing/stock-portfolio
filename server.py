"""
OneDrive 엑셀 데이터 분석 서버
- OneDrive 로컬 동기화 폴더에서 엑셀 파일을 읽어 JSON API로 제공
- Flask 기반 REST API 서버
"""

import os
import sys
import json
import re
from pathlib import Path

# 작업 디렉토리를 스크립트 위치로 고정 (작업 스케줄러 실행 시 System32 참조 방지)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(BASE_DIR)
from flask import Flask, jsonify, request, send_from_directory, make_response
from flask_cors import CORS
import pandas as pd
import openpyxl
import subprocess
import threading
from openpyxl.styles import Font

# Windows 콘솔 CP949 인코딩 충돌 방지 - stdout/stderr를 UTF-8로 강제 설정 (백그라운드 예외 처리)
try:
    if sys.stdout and hasattr(sys.stdout, 'encoding') and sys.stdout.encoding != 'utf-8':
        try:
            sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)
        except Exception:
            pass
    if sys.stderr and hasattr(sys.stderr, 'encoding') and sys.stderr.encoding != 'utf-8':
        try:
            sys.stderr = open(sys.stderr.fileno(), mode='w', encoding='utf-8', buffering=1)
        except Exception:
            pass
except Exception:
    pass


app = Flask(__name__, static_folder=None)
# 모든 경로와 오리진에 대해 CORS 허용 (모바일 앱 접속용)
CORS(app, resources={r"/*": {"origins": "*"}})

@app.before_request
def handle_options_preflight():
    if request.method == 'OPTIONS':
        response = make_response()
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
        response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
        return response

@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

# API 응답 gzip 압축 (대용량 JSON 전송 최적화 - 약 70~80% 크기 감소)
try:
    from flask_compress import Compress
    compress = Compress()
    app.config['COMPRESS_MIMETYPES'] = [
        'application/json',
        'text/html',
        'text/css',
        'application/javascript',
    ]
    app.config['COMPRESS_LEVEL'] = 6      # 압축 레벨 (1~9, 6이 속도/크기 균형)
    app.config['COMPRESS_MIN_SIZE'] = 500 # 500바이트 이상만 압축
    compress.init_app(app)
    print("[COMPRESS] gzip 압축 활성화됨")
except ImportError:
    print("[COMPRESS] flask-compress 미설치 — pip install flask-compress 로 설치 가능")


# 글로벌 OneDrive 경로 설정 (사용자 지정 경로 우선)
def find_target_onedrive_path():
    fixed_path = r"C:\Users\zerod\OneDrive"
    if os.path.isdir(fixed_path):
        return fixed_path
    
    # 자동 탐색 로직 (예비용)
    possible_paths = [
        os.path.expanduser("~/OneDrive"),
        os.path.expanduser("~/OneDrive - Personal"),
        os.path.expandvars(r"%USERPROFILE%\OneDrive"),
    ]
    for path in possible_paths:
        if path and os.path.isdir(path):
            return os.path.abspath(path)
    return None

ONEDRIVE_PATH = find_target_onedrive_path()

# ==================== 자동 동기화 트리거 ====================
_export_lock = threading.Lock()

def trigger_export():
    """데이터 변경 시 export_to_json.py를 실행하여 앱용 JSON 갱신 (비동기)"""
    def run():
        # 이미 변환 중이면 새 요청 무시 (충돌 및 IO 과부하 방지)
        if not _export_lock.acquire(blocking=False):
            print("[AUTO-SYNC] 이미 JSON 변환 중입니다. 이번 요청은 건너뜁니다.")
            return
        try:
            print("[AUTO-SYNC] JSON 변환 시작...")
            subprocess.run([sys.executable, "export_to_json.py"], check=True)
            print("[AUTO-SYNC] ✅ JSON 변환 완료!")
        except Exception as e:
            print(f"[AUTO-SYNC] ❌ 변환 실패: {e}")
        finally:
            _export_lock.release()

    # 서버 응답을 방해하지 않도록 별도 스레드에서 실행
    threading.Thread(target=run).start()


def git_has_changes():
    """Git 작업 디렉터리에 커밋되지 않은 변경 사항이 있는지 확인 (로캘에 구애받지 않음)"""
    try:
        res = subprocess.run(
            ["git", "status", "--porcelain"],
            cwd=BASE_DIR, capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=15
        )
        return len(res.stdout.strip()) > 0
    except Exception as e:
        print(f"[GIT-CHECK] 변경사항 확인 오류: {e}")
        return True  # 오류 발생 시 보수적으로 변경사항이 있는 것으로 취급

def trigger_export_and_push_sync():
    """동기적으로 JSON 내보내기 + Git Push까지 실행 (sync-receive 전용)
    
    모바일에서 PC로 전송 후 리다이렉트 전에 호출하여,
    GitHub Pages에 최신 데이터가 반영된 상태에서 앱이 데이터를 받을 수 있게 합니다.
    """
    import time
    lock_file = os.path.join(BASE_DIR, ".git_sync.lock")
    
    # 1. 락 획득 시도 (최대 15초 대기)
    for i in range(15):
        if not os.path.exists(lock_file):
            break
        print(f"[SYNC-PUSH] ⚠️ Git 동기화 락 감지. 대기 중... ({i+1}/15)")
        time.sleep(1)
        
    # 락 파일 생성
    try:
        with open(lock_file, "w", encoding="utf-8") as f:
            f.write(str(os.getpid()))
    except Exception as e:
        print(f"[SYNC-PUSH] 락 파일 생성 실패: {e}")

    with _export_lock:
        try:
            # 1단계: JSON 내보내기
            print("[SYNC-PUSH] 1. JSON 내보내기 시작...")
            result = subprocess.run(
                [sys.executable, "export_to_json.py"],
                capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=180
            )
            if result.returncode != 0:
                print(f"[SYNC-PUSH] ❌ JSON 내보내기 실패: {result.stderr[:300]}")
                return False
            print("[SYNC-PUSH] ✅ JSON 내보내기 완료")

            # 1-5단계: 신호 데이터(이평선/RSI) 내보내기 (텍스트 동기화 시에는 40~50초 소요되어 응답 지연을 방지하기 위해 주석 처리)
            # print("[SYNC-PUSH] 1-5. 신호 데이터(이평선/RSI) 내보내기 시작...")
            # result_sig = subprocess.run(
            #     [sys.executable, "export_signals.py"],
            #     capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=300
            # )
            # if result_sig.returncode != 0:
            #     print(f"[SYNC-PUSH] ❌ 신호 데이터 내보내기 실패: {result_sig.stderr[:300]}")
            # else:
            #     print("[SYNC-PUSH] ✅ 신호 데이터 내보내기 완료")

            # 2단계: 모바일 데이터 복사 (StockPortfolioApp/public/data → mobile/data)
            print("[SYNC-PUSH] 2. 모바일 데이터 복사 중...")
            src = os.path.join(BASE_DIR, "StockPortfolioApp", "public", "data")
            dst = os.path.join(BASE_DIR, "mobile", "data")
            if os.path.exists(src):
                subprocess.run(
                    f'xcopy "{src}" "{dst}" /E /I /Y',
                    shell=True, capture_output=True, timeout=30
                )
            print("[SYNC-PUSH] ✅ 모바일 데이터 복사 완료")

            # 3단계: Git add + commit + push
            print("[SYNC-PUSH] 3. Git push 시작...")
            subprocess.run(["git", "add", "."], cwd=BASE_DIR, capture_output=True, timeout=30)
            
            # 변경 사항이 없는 경우 커밋 건너뜀
            if not git_has_changes():
                print("[SYNC-PUSH] 변경사항 없음, 커밋 및 푸시 건너뜀")
                return True

            from datetime import datetime as dt
            commit_msg = f"Mobile sync update {dt.now().strftime('%Y-%m-%d %H:%M')}"
            commit_result = subprocess.run(
                ["git", "commit", "-m", commit_msg],
                cwd=BASE_DIR, capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=30
            )
            if commit_result.returncode != 0:
                print(f"[SYNC-PUSH] ❌ Git 커밋 실패: {commit_result.stderr[:300]}")
                return False

            env = os.environ.copy()
            env["GCM_INTERACTIVE"] = "never"
            env["GIT_TERMINAL_PROMPT"] = "0"
            
            # push 직전 rebase pull로 충돌 방지
            print("[SYNC-PUSH] push 전 원격 변경 사항 병합 (rebase)...")
            subprocess.run(
                ["git", "pull", "--rebase", "origin", "main"],
                cwd=BASE_DIR, capture_output=True, env=env, timeout=60
            )

            push_result = subprocess.run(
                ["git", "push", "origin", "main"],
                cwd=BASE_DIR, capture_output=True, text=True, encoding="utf-8", errors="replace",
                timeout=120, env=env
            )
            if push_result.returncode == 0:
                print("[SYNC-PUSH] ✅ GitHub push 완료!")
                return True
            else:
                print(f"[SYNC-PUSH] ❌ GitHub push 실패: {push_result.stderr[:300]}")
                return False

        except subprocess.TimeoutExpired as e:
            print(f"[SYNC-PUSH] ❌ 시간 초과: {e}")
            return False
        except Exception as e:
            print(f"[SYNC-PUSH] ❌ 예외 발생: {e}")
            return False
        finally:
            # 락 해제
            if os.path.exists(lock_file):
                try:
                    os.remove(lock_file)
                except:
                    pass

# ==================== 유틸리티 함수 (취소선 처리) ====================
def parse_strikethrough_text(text):
    """텍스트 내의 <del>태그 또는 ~~패턴을 파싱하여 CellRichText 객체로 변환"""
    if not isinstance(text, str) or not text:
        return text
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    
    text = re.sub(r'<del>(.*?)</del>', r'~~\1~~', text)
    pattern = r'~~(.*?)~~'
    parts = re.split(pattern, text)
    
    if len(parts) == 1: return text
    
    rich_text = CellRichText()
    default_font = InlineFont(rFont='맑은 고딕')
    strike_font = InlineFont(rFont='맑은 고딕', strike=True)
    
    for i, part in enumerate(parts):
        if not part: continue
        if i % 2 == 0:
            rich_text.append(TextBlock(default_font, part))
        else:
            rich_text.append(TextBlock(strike_font, part))
    return rich_text

def extract_rich_text(cell):
    """엑셀 셀의 실제 취소선 서식을 감지하여 마크다운(~~) 형식으로 변환"""
    from openpyxl.cell.rich_text import CellRichText
    if not hasattr(cell, 'value') or cell.value is None: return ""
    if isinstance(cell.value, CellRichText):
        result = []
        for part in cell.value:
            text = part.text if hasattr(part, 'text') else str(part)
            if hasattr(part, 'font') and part.font and part.font.strike:
                result.append(f"~~{text}~~")
            else: result.append(text)
        return "".join(result)
    if cell.font and cell.font.strike: return f"~~{cell.value}~~"
    return str(cell.value)

@app.route('/')
def index():
    """메인 페이지 서빙"""
    return send_from_directory('.', 'index.html')

@app.route('/mobile')
@app.route('/mobile/')
def mobile_index():
    """모바일 PWA 메인 서빙"""
    return send_from_directory('mobile', 'index.html')



@app.route('/mobile/data/<path:filename>', methods=['GET', 'OPTIONS'])
def serve_mobile_data(filename):
    if request.method == 'OPTIONS':
        response = make_response()
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
        response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
        return response
    
    file_path = os.path.join('mobile/data', filename)
    if not os.path.exists(file_path):
        return jsonify({'error': 'File not found'}), 404
        
    try:
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            content = f.read()
        response = make_response(content)
        response.headers['Content-Type'] = 'application/json; charset=utf-8'
        response.headers['Access-Control-Allow-Origin'] = '*'
        return response
    except Exception as e:
        print(f"[CORS-SERVE] 모바일 데이터 로드 오류 ({filename}): {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/StockPortfolioApp/public/data/<path:filename>', methods=['GET', 'OPTIONS'])
def serve_app_data(filename):
    if request.method == 'OPTIONS':
        response = make_response()
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
        response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
        return response
        
    file_path = os.path.join('StockPortfolioApp/public/data', filename)
    if not os.path.exists(file_path):
        return jsonify({'error': 'File not found'}), 404
        
    try:
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            content = f.read()
        response = make_response(content)
        response.headers['Content-Type'] = 'application/json; charset=utf-8'
        response.headers['Access-Control-Allow-Origin'] = '*'
        return response
    except Exception as e:
        print(f"[CORS-SERVE] 앱 데이터 로드 오류 ({filename}): {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/onedrive-status')
def onedrive_status():
    """OneDrive 연결 상태 확인"""
    if ONEDRIVE_PATH and os.path.isdir(ONEDRIVE_PATH):
        return jsonify({
            'connected': True,
            'path': ONEDRIVE_PATH,
            'message': f'OneDrive 연결됨: {ONEDRIVE_PATH}'
        })
    return jsonify({
        'connected': False,
        'path': None,
        'message': 'OneDrive 동기화 폴더를 찾을 수 없습니다. 경로를 직접 지정해주세요.'
    })

@app.route('/api/set-path', methods=['POST'])
def set_path():
    """OneDrive 경로 수동 설정"""
    global ONEDRIVE_PATH
    data = request.get_json()
    path = data.get('path', '')
    
    if os.path.isdir(path):
        ONEDRIVE_PATH = path
        return jsonify({'success': True, 'path': ONEDRIVE_PATH})
    return jsonify({'success': False, 'message': '유효하지 않은 경로입니다.'}), 400

@app.route('/api/files')
def list_excel_files():
    """OneDrive 내 엑셀 파일 목록 조회"""
    if not ONEDRIVE_PATH:
        return jsonify({'error': 'OneDrive 경로가 설정되지 않았습니다.'}), 400
    
    search_path = request.args.get('subdir', '')
    base_path = os.path.join(ONEDRIVE_PATH, search_path)
    
    if not os.path.isdir(base_path):
        return jsonify({'error': f'경로를 찾을 수 없습니다: {base_path}'}), 404
    
    files = []
    # 현재 폴더의 하위 디렉토리 목록
    directories = []
    
    try:
        for item in os.listdir(base_path):
            full_path = os.path.join(base_path, item)
            rel_path = os.path.relpath(full_path, ONEDRIVE_PATH)
            
            if os.path.isdir(full_path):
                # 숨김 폴더 제외
                if not item.startswith('.'):
                    directories.append({
                        'name': item,
                        'path': rel_path.replace('\\', '/')
                    })
            elif item.lower().endswith(('.xlsx', '.xls', '.xlsm')):
                stat = os.stat(full_path)
                files.append({
                    'name': item,
                    'path': rel_path.replace('\\', '/'),
                    'size': stat.st_size,
                    'modified': stat.st_mtime,
                    'full_path': full_path
                })
    except PermissionError:
        return jsonify({'error': '해당 폴더에 대한 접근 권한이 없습니다.'}), 403
    
    return jsonify({
        'current_dir': search_path or '/',
        'directories': sorted(directories, key=lambda x: x['name']),
        'files': sorted(files, key=lambda x: x['name'])
    })

# 파일 파싱 캐시
EXCEL_CACHE = {}

@app.route('/api/read-excel')
def read_excel():
    """엑셀 파일 데이터 읽기"""
    if not ONEDRIVE_PATH:
        return jsonify({'error': 'OneDrive 경로가 설정되지 않았습니다.'}), 400
    
    file_path = request.args.get('file', '')
    sheet_name = request.args.get('sheet', None)
    
    full_path = os.path.join(ONEDRIVE_PATH, file_path)
    
    if not os.path.isfile(full_path):
        return jsonify({'error': f'파일을 찾을 수 없습니다: {file_path}'}), 404
    
    try:
        import io
        
        # 캐시 확인 (수정 시간이 같으면 캐시 반환)
        file_stat = os.stat(full_path)
        last_modified = file_stat.st_mtime
        
        cache_key = f"{full_path}_{sheet_name}_{last_modified}"
        if cache_key in EXCEL_CACHE:
            return jsonify(EXCEL_CACHE[cache_key])
            
        # 파일 잠금(Lock)을 방지하기 위해 메모리로 먼저 읽어오기
        with open(full_path, 'rb') as f:
            file_data = f.read()
            
        # 시트 목록 가져오기
        xl = pd.ExcelFile(io.BytesIO(file_data), engine='openpyxl')
        sheet_names = xl.sheet_names
        
        # 특정 시트 또는 첫 번째 시트 읽기
        target_sheet = sheet_name if sheet_name else sheet_names[0]
        df = pd.read_excel(xl, sheet_name=target_sheet)
        
        # NaN 값 처리
        df = df.fillna('')

        # --- 실적 시트 특수 처리 (헤더 자동 탐색) ---
        header_row_idx = 0
        if "실적" in target_sheet and not df.empty:
            # 첫 10행 내에서 '연도' 또는 '수익율' 키워드 찾기
            found_header = False
            for i in range(min(10, len(df))):
                row_vals = [str(x).strip() for x in df.iloc[i].values]
                if any("연도" in val or "수익율" in val for val in row_vals):
                    header_row_idx = i
                    # 현재 행을 컬럼명으로 승격
                    new_cols = []
                    for j, val in enumerate(row_vals):
                        if val and val != 'nan':
                            new_cols.append(val)
                        else:
                            new_cols.append(f"Unnamed: {j}")
                    df.columns = new_cols
                    df = df.iloc[i+1:].reset_index(drop=True)
                    found_header = True
                    break
            
            # 만약 못 찾았더라도 '연도' 컬럼이 B열(1번 인덱스)에 있는 경우가 많으므로 보정
            if not found_header and len(df.columns) > 1:
                # B4가 15년이면 B3가 헤더일 가능성 높음 (pandas 0-indexed 기준 Row 2)
                pass

        # 숫자형 컬럼 감지 (강제 변환 시도 포함)
        for col in df.columns:
            if '연도' in str(col): continue 
            try:
                temp_numeric = pd.to_numeric(df[col].replace('', pd.NA), errors='coerce')
                if temp_numeric.notna().any():
                    if temp_numeric.notna().sum() / len(df) > 0.2: # 기준 완화
                        df[col] = temp_numeric.fillna(0)
            except:
                pass

        numeric_columns = df.select_dtypes(include=['number']).columns.tolist()
        
        # 기본 통계 계산
        stats = {}
        for col in numeric_columns:
            col_data = df[col].replace('', pd.NA).dropna()
            if len(col_data) > 0:
                stats[col] = {
                    'mean': round(float(col_data.mean()), 2),
                    'sum': round(float(col_data.sum()), 2),
                    'min': round(float(col_data.min()), 2),
                    'max': round(float(col_data.max()), 2),
                    'count': int(col_data.count()),
                    'std': round(float(col_data.std()), 2) if len(col_data) > 1 else 0
                }
        
        # 데이터를 JSON 직렬화 가능한 형태로 변환
        columns = [str(c) for c in df.columns.tolist()]
        
        # 탐구생활/실적 시트인 경우 서식 데이터(취소선 등)를 위해 openpyxl로 재확인
        is_special = any(k in target_sheet for k in ["탐구", "실적"])
        if is_special:
            # openpyxl 워크북 로드
            wb_format = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True, rich_text=True)
            ws_format = wb_format[target_sheet]
            
            data = []
            # openpyxl은 1-based index이므로 pandas 인덱스 기반 보정
            # pandas df.iloc[header_row_idx]는 원본 엑셀의 row=header_row_idx+2 (기본 헤더가 1행일 때)
            # 하지만 df를 처음 읽을 때 이미 Row 1이 헤더로 쓰였으므로:
            # Row 1 (Header), df.iloc[0] = Row 2, df.iloc[1] = Row 3 ...
            # 따라서 header_row_idx일 때 엑셀 행은 header_row_idx + 2
            start_row = header_row_idx + 3 if "실적" in target_sheet else 2
            
            for r_idx in range(start_row, ws_format.max_row + 1):
                row_data = {'_realIndex': r_idx - 2}
                for c_idx, col_name in enumerate(columns, start=1):
                    cell = ws_format.cell(row=r_idx, column=c_idx)
                    row_data[col_name] = extract_rich_text(cell)
                data.append(row_data)
        else:
            data = []
            for idx, row in df.iterrows():
                row_data = {'_realIndex': idx}
                for i, col in enumerate(df.columns):
                    val = row[col]
                    col_name = columns[i]
                    if pd.isna(val): row_data[col_name] = ''
                    elif isinstance(val, (int, float)): row_data[col_name] = val
                    else: row_data[col_name] = str(val)
                data.append(row_data)
        
        # 파일 수정 시간 가져오기
        file_stat = os.stat(full_path)
        last_modified = file_stat.st_mtime

        response_data = {
            'file_name': os.path.basename(file_path),
            'last_modified': last_modified,
            'sheet_names': sheet_names,
            'current_sheet': target_sheet,
            'columns': columns,
            'numeric_columns': [str(c) for c in numeric_columns],
            'data': data,
            'row_count': len(data),
            'stats': stats
        }
        
        # 캐시 저장 (메모리 누수 방지를 위해 10개 초과 시에만 정리)
        if len(EXCEL_CACHE) > 10:
            EXCEL_CACHE.clear()
        EXCEL_CACHE[cache_key] = response_data

        return jsonify(response_data)
    
    except Exception as e:
        return jsonify({'error': f'엑셀 파일 읽기 오류: {str(e)}'}), 500

@app.route('/api/analyze')
def analyze_data():
    """엑셀 데이터 심층 분석"""
    if not ONEDRIVE_PATH:
        return jsonify({'error': 'OneDrive 경로가 설정되지 않았습니다.'}), 400
    
    file_path = request.args.get('file', '')
    sheet_name = request.args.get('sheet', None)
    
    full_path = os.path.join(ONEDRIVE_PATH, file_path)
    
    if not os.path.isfile(full_path):
        return jsonify({'error': f'파일을 찾을 수 없습니다: {file_path}'}), 404
    
    try:
        import io
        # 파일 잠금(Lock)을 방지하기 위해 메모리로 먼저 읽어오기
        with open(full_path, 'rb') as f:
            file_data = f.read()
            
        df = pd.read_excel(io.BytesIO(file_data), sheet_name=sheet_name, engine='openpyxl')
        df = df.fillna('')
        
        # 컬럼 타입 분석
        column_analysis = []
        for col in df.columns:
            col_data = df[col].replace('', pd.NA).dropna()
            col_info = {
                'name': str(col),
                'total': len(df),
                'non_empty': int(col_data.count()) if hasattr(col_data, 'count') else len(col_data),
                'unique': int(df[col].nunique()),
            }
            
            # 숫자형인지 확인
            numeric_data = pd.to_numeric(col_data, errors='coerce').dropna()
            if len(numeric_data) > 0 and len(numeric_data) / max(len(col_data), 1) > 0.5:
                col_info['type'] = 'numeric'
                col_info['stats'] = {
                    'mean': round(float(numeric_data.mean()), 2),
                    'median': round(float(numeric_data.median()), 2),
                    'sum': round(float(numeric_data.sum()), 2),
                    'min': round(float(numeric_data.min()), 2),
                    'max': round(float(numeric_data.max()), 2),
                    'std': round(float(numeric_data.std()), 2) if len(numeric_data) > 1 else 0,
                }
            else:
                col_info['type'] = 'text'
                # 상위 빈도 값
                value_counts = df[col].value_counts().head(10)
                col_info['top_values'] = [
                    {'value': str(v), 'count': int(c)} 
                    for v, c in value_counts.items() if str(v).strip()
                ]
            
            column_analysis.append(col_info)
        
        return jsonify({
            'file_name': os.path.basename(file_path),
            'total_rows': len(df),
            'total_columns': len(df.columns),
            'column_analysis': column_analysis
        })
    
    except Exception as e:
        return jsonify({'error': f'분석 오류: {str(e)}'}), 500


@app.route('/api/save-journal', methods=['POST'])
def save_journal():
    """매매일지 데이터 추가 저장"""
    if not ONEDRIVE_PATH:
        return jsonify({'error': 'OneDrive 경로가 설정되지 않았습니다.'}), 400
    
    data = request.get_json()
    file_path = data.get('file', '')
    sheet_name = data.get('sheet', '매매일지')
    row_values = data.get('row', [])  # [날짜, 종목, 수량, 단가, 매매종류, 투자금]
    
    full_path = os.path.join(ONEDRIVE_PATH, file_path)
    
    if not os.path.isfile(full_path):
        return jsonify({'error': f'파일을 찾을 수 없습니다: {file_path}'}), 404
    
    try:
        from openpyxl.styles import Font, PatternFill

        # 포트폴리오 맵 업데이트에 사용할 스타일 정의
        yellow_fill = PatternFill(
            fill_type="solid",
            fgColor="FFFF00"   # 노란색
        )
        no_fill = PatternFill(fill_type=None)  # 배경색 없음(초기화용)

        # openpyxl을 사용하여 데이터 추가
        wb = openpyxl.load_workbook(full_path, rich_text=True)
        try:
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
            else:
                ws = wb[sheet_name]
            
            # 데이터 추가
            ws.append(row_values)
            
            # 마지막으로 추가된 행의 폰트 색상 설정
            last_row = ws.max_row
            trade_type_val = row_values[4] if len(row_values) > 4 else ""
            
            # 색상 결정 (매도: 빨간색, 매수: 검정색)
            font_color = "FF0000" if trade_type_val == "매도" else "000000"
            cell_font = Font(color=font_color)
            
            for col_idx in range(1, len(row_values) + 1):
                ws.cell(row=last_row, column=col_idx).font = cell_font
            
            # --- 포트폴리오 맵 절대값 동기화 ---
            # row_values: [날짜, 종목, 수량, 단가, 매매종류, 투자금]
            try:
                trade_stock = str(row_values[1]).strip() if len(row_values) > 1 else ""
                # 투자금액 절대값 계산
                raw_investment = float(row_values[5]) if len(row_values) > 5 else 0
                investment_amount = abs(raw_investment)
                
                if trade_stock:
                    sync_portfolio_map(wb, trade_stock, investment_amount)
                    
                    # 실제 반영된 마크 수 재검증 → 매매일지 투자금 보정
                    ws_map = wb['포트폴리오 맵']
                    trade_stock_clean = trade_stock.replace(" ", "")
                    final_ones = 0
                    for r in range(1, ws_map.max_row + 1):
                        cell_val = str(ws_map.cell(row=r, column=4).value or "").strip().replace(" ", "")
                        if cell_val == trade_stock_clean:
                            for c in range(5, 101):
                                if ws_map.cell(row=r, column=c).value == 1:
                                    final_ones += 1
                            ws.cell(row=last_row, column=6).value = final_ones * 100
                            break
                    print(f"[OK] [{trade_stock}] sync done: marks={final_ones}")
            except Exception as inner_e:
                print(f"[ERROR] portfolio map update failed: {inner_e}")
            
            # 최종 저장 (매매일지 + 포트폴리오 맵 모두 반영)
            wb.save(full_path)
            wb.close()
            
            # 아이폰 앱용 데이터 자동 갱신
            trigger_export()
                
        finally:
            # 오류가 발생하더라도 확실하게 파일을 닫아 잠금 해제
            try: wb.close()
            except: pass
        
        return jsonify({'success': True, 'message': '성공적으로 저장되었습니다.'})
    
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': f'저장 오류: {str(e)}'}), 500


@app.route('/api/update-row', methods=['POST'])
def update_row():
    if not ONEDRIVE_PATH:
        return jsonify({'error': 'OneDrive 경로가 설정되지 않았습니다.'}), 400

    data = request.get_json()
    file_path = data.get('file', '')
    sheet_name = data.get('sheet')
    row_index = int(data.get('rowIndex', 0))
    values = data.get('values', [])

    if not sheet_name:
        return jsonify({'error': 'sheet 값이 필요합니다.'}), 400

    full_path = os.path.join(ONEDRIVE_PATH, file_path)
    if not os.path.isfile(full_path):
        return jsonify({'error': f'파일을 찾을 수 없습니다: {file_path}'}), 404

    try:
        from openpyxl.styles import Alignment
        wb = openpyxl.load_workbook(full_path, rich_text=True)
        try:
            if sheet_name not in wb.sheetnames:
                return jsonify({'error': f'시트를 찾을 수 없습니다: {sheet_name}'}), 404

            ws = wb[sheet_name]
            target_row = row_index + 2
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=target_row, column=col_idx)
                # 취소선 처리
                processed_value = parse_strikethrough_text(value)
                cell.value = processed_value
                
                # 줄바꿈(\n)이 포함된 셀은 엑셀에서도 표시되도록 wrap_text 설정
                if isinstance(value, str) and '\n' in value:
                    cell.alignment = Alignment(wrap_text=True)

            # 매매일지인 경우 포트폴리오 맵 동기화
            if sheet_name == '매매일지':
                try:
                    stock_name = values[1] if len(values) > 1 else ""
                    amount = float(values[5]) if len(values) > 5 else 0
                    if stock_name:
                        sync_portfolio_map(wb, stock_name, amount)
                except: pass

            wb.save(full_path)
        finally:
            try: wb.close()
            except: pass
        
        # 아이폰 앱용 데이터 자동 갱신
        trigger_export()
        
        return jsonify({'success': True, 'message': '행이 업데이트되었습니다.'})
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': f'업데이트 오류: {str(e)}'}), 500

@app.route('/api/sync-receive', methods=['GET', 'POST'])
def sync_receive():
    if not ONEDRIVE_PATH:
        return "OneDrive 경로가 설정되지 않았습니다.", 400

    try:
        import json
        if request.is_json:
            edits = request.get_json()
        else:
            payload = request.values.get('payload')
            if not payload:
                return "No payload provided.", 400
            edits = json.loads(payload)

        if not isinstance(edits, list):
            edits = [edits]

        from openpyxl.styles import Alignment
        import traceback
        from collections import defaultdict

        # ★ 파일별로 edit을 그룹화하여 동일 파일은 1번만 열기 (성능 최적화)
        edits_by_file = defaultdict(list)
        for edit in edits:
            file_path = edit.get('file', '')
            if os.path.isabs(file_path) and os.path.isfile(file_path):
                full_path = file_path
            else:
                full_path = os.path.join(ONEDRIVE_PATH, file_path)
            edits_by_file[full_path].append(edit)

        for full_path, file_edits in edits_by_file.items():
            if not os.path.isfile(full_path):
                print(f'[sync-receive] 파일 없음: {full_path}')
                continue

            wb = openpyxl.load_workbook(full_path, rich_text=True)

            try:
                for edit in file_edits:
                    sheet_name = edit.get('sheet')
                    row_index = int(edit.get('rowIndex', 0))
                    values = edit.get('values', [])

                    if sheet_name not in wb.sheetnames:
                        continue

                    ws = wb[sheet_name]
                    # 기본값: pandas iloc[N] → openpyxl row = N+2
                    target_row = row_index + 2
                    
                    stock_name = edit.get('stockName', '').strip().replace(' ', '')
                    if stock_name:
                        # 1. 헤더에서 '종목명' 컬럼 인덱스 찾기 (보통 1~3행 사이)
                        name_col_idx = None
                        for r in range(1, 4):
                            for c in range(1, ws.max_column + 1):
                                val = str(ws.cell(row=r, column=c).value or '').strip()
                                if val == '종목명' or val == 'Unnamed: 1':
                                    name_col_idx = c
                                    break
                            if name_col_idx:
                                break
                        
                        # 2. 헤더를 못 찾았으면 기본 B열(2)로 가정
                        if not name_col_idx:
                            name_col_idx = 2
                            
                        # 3. 해당 컬럼에서 종목명이 일치하는 행 찾기
                        found_row = None
                        for r in range(1, ws.max_row + 1):
                            cell_val = str(ws.cell(row=r, column=name_col_idx).value or '').strip().replace(' ', '')
                            if cell_val and cell_val == stock_name:
                                found_row = r
                                break
                        
                        if found_row:
                            target_row = found_row
                            print(f'[sync-receive] 종목명 "{stock_name}" 매칭 성공! -> 엑셀 {target_row}행 덮어쓰기 진행')
                        else:
                            print(f'[sync-receive] 종목명 "{stock_name}" 매칭 실패! -> 기존 로직대로 {target_row}행에 덮어씁니다.')

                    for col_idx, value in enumerate(values, start=1):
                        cell = ws.cell(row=target_row, column=col_idx)
                        processed_value = parse_strikethrough_text(value)
                        cell.value = processed_value
                        if isinstance(value, str) and '\n' in value:
                            cell.alignment = Alignment(wrap_text=True)

                # 파일당 1회만 저장/닫기
                wb.save(full_path)
            finally:
                try: wb.close()
                except: pass

        # ★ 모바일 동기화: JSON 내보내기 + Git Push를 동기적으로 실행
        # 비동기 trigger_export() 대신 동기적으로 실행하여
        # GitHub Pages에 최신 데이터가 반영된 후 리다이렉트합니다.
        push_success = trigger_export_and_push_sync()

        if push_success:
            return """
            <html>
            <head>
              <meta name="viewport" content="width=device-width, initial-scale=1.0">
              <script>
                // GitHub Pages CDN 반영 대기 (약 5초) 후 리다이렉트
                setTimeout(function() {
                  window.location.href = 'https://smuth-swing.github.io/stock-portfolio/mobile/?sync=success';
                }, 5000);
              </script>
              <style>
                body { background: #0F172A; color: white; font-family: sans-serif; text-align: center; padding-top: 100px; }
                h2 { color: #00F2FE; }
                p { color: #94A3B8; font-size: 16px; margin-top: 20px; }
                .spinner { display: inline-block; width: 40px; height: 40px; border: 4px solid rgba(0,242,254,0.2); border-top: 4px solid #00F2FE; border-radius: 50%; animation: spin 1s linear infinite; margin-top: 30px; }
                @keyframes spin { to { transform: rotate(360deg); } }
                button { background: #00F2FE; color: #0F172A; border: none; padding: 15px 30px; border-radius: 10px; font-size: 18px; font-weight: bold; margin-top: 30px; cursor: pointer; }
              </style>
            </head>
            <body>
              <h2>✅ PC 서버 반영 완료!</h2>
              <p>GitHub Pages 반영 대기 중... (약 5초)</p>
              <div class="spinner"></div>
              <br>
              <button onclick="window.location.href='https://smuth-swing.github.io/stock-portfolio/mobile/?sync=success'">바로 돌아가기</button>
            </body>
            </html>
            """
        else:
            # Push 실패 시에도 엑셀에는 이미 저장되었으므로, 큐 정리를 위해 success로 리다이렉트
            return """
            <html>
            <head>
              <meta name="viewport" content="width=device-width, initial-scale=1.0">
              <script>
                setTimeout(function() {
                  window.location.href = 'https://smuth-swing.github.io/stock-portfolio/mobile/?sync=success';
                }, 3000);
              </script>
              <style>
                body { background: #0F172A; color: white; font-family: sans-serif; text-align: center; padding-top: 100px; }
                h2 { color: #EAB308; }
                p { color: #94A3B8; font-size: 14px; margin-top: 15px; }
                button { background: #EAB308; color: #422006; border: none; padding: 15px 30px; border-radius: 10px; font-size: 18px; font-weight: bold; margin-top: 30px; cursor: pointer; }
              </style>
            </head>
            <body>
              <h2>⚠️ PC에 저장됨 (GitHub 반영은 잠시 후)</h2>
              <p>엑셀 파일에는 정상 저장되었습니다.</p>
              <p>GitHub Pages 반영은 자동 업로더가 곧 처리합니다.</p>
              <button onclick="window.location.href='https://smuth-swing.github.io/stock-portfolio/mobile/?sync=success'">돌아가기</button>
            </body>
            </html>
            """
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return f"동기화 오류: {str(e)}", 500


def sync_portfolio_map(wb, stock_name, trade_amount):
    """포트폴리오 맵의 종목 마크(점)를 투자금액에 맞게 동기화"""
    if '포트폴리오 맵' not in wb.sheetnames:
        return
    
    from openpyxl.styles import PatternFill
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    no_fill = PatternFill(fill_type=None)
    
    ws_map = wb['포트폴리오 맵']
    stock_clean = stock_name.replace(" ", "")
    target_marks = int(abs(trade_amount) // 100)
    
    # 종목명이 일치하는 모든 행을 찾아서 처리
    for r in range(1, ws_map.max_row + 1):
        cell_val = str(ws_map.cell(row=r, column=4).value or "").strip().replace(" ", "")
        if cell_val and (stock_clean in cell_val or cell_val in stock_clean):
            # 1. 초기화
            for c in range(5, 101):
                cell = ws_map.cell(row=r, column=c)
                cell.value = None
                cell.fill = no_fill
            # 2. 마킹
            if target_marks > 0:
                for i in range(target_marks):
                    col_idx = 5 + i
                    if col_idx <= 100:
                        cell = ws_map.cell(row=r, column=col_idx)
                        cell.value = 1
                        cell.fill = yellow_fill
    print(f"📊 [{stock_name}] 포트폴리오 동기화: {target_marks}개")


@app.route('/api/delete-row', methods=['POST'])
def delete_row():
    """행 삭제 API - 매매일지의 경우 이전 데이터로 동기화 포함"""
    if not ONEDRIVE_PATH:
        return jsonify({'error': 'OneDrive 경로가 설정되지 않았습니다.'}), 400

    data = request.get_json()
    file_path = data.get('file', '')
    sheet_name = data.get('sheet')
    row_index = int(data.get('rowIndex', 0))

    full_path = os.path.join(ONEDRIVE_PATH, file_path)
    if not os.path.isfile(full_path):
        return jsonify({'error': f'파일을 찾을 수 없습니다: {file_path}'}), 404

    try:
        wb = openpyxl.load_workbook(full_path, rich_text=True)
        try:
            if sheet_name not in wb.sheetnames:
                return jsonify({'error': f'시트를 찾을 수 없습니다: {sheet_name}'}), 404

            ws = wb[sheet_name]
            target_row_idx = row_index + 2
            
            # 삭제 전 정보 기억
            stock_name = None
            if sheet_name == '매매일지':
                stock_name = str(ws.cell(row=target_row_idx, column=2).value or "").strip()
            
            # 행 삭제
            ws.delete_rows(target_row_idx)
            
            # 매매일지 삭제 후 동기화
            if sheet_name == '매매일지' and stock_name:
                last_amount = 0
                stock_clean = stock_name.replace(" ", "")
                for r in range(2, ws.max_row + 1):
                    cell_val = str(ws.cell(row=r, column=2).value or "").strip().replace(" ", "")
                    if cell_val == stock_clean:
                        try:
                            val = ws.cell(row=r, column=6).value
                            if val is not None: last_amount = float(val)
                        except: pass
                
                sync_portfolio_map(wb, stock_name, last_amount)
                print(f"🗑️ [{stock_name}] 삭제 및 이전 데이터({last_amount}) 동기화")

            wb.save(full_path)
        finally:
            try: wb.close()
            except: pass
        
        # 아이폰 앱용 데이터 자동 갱신
        trigger_export()
        
        return jsonify({'success': True, 'message': '행이 삭제되었습니다.'})
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': f'삭제 오류: {str(e)}'}), 500


@app.route('/api/ping')
def ping():
    """
    절전 복귀 감지용 헬스체크 엔드포인트.
    클라이언트(PWA)가 주기적으로 호출하여 서버 생존 및 절전 복귀를 감지한다.
    """
    import time
    return jsonify({
        'alive': True,
        'timestamp': time.time(),
        'server_time': __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })


# ══════════════════════════════════════════════════════════════
# LS증권 OpenAPI 연동 엔드포인트
# ══════════════════════════════════════════════════════════════

@app.route('/api/ls/config', methods=['GET', 'POST'])
def ls_config():
    """LS증권 API 설정 저장 / 조회"""
    try:
        from ls_api import load_config, save_config
    except ImportError:
        return jsonify({'error': 'ls_api 모듈을 찾을 수 없습니다.'}), 500

    if request.method == 'GET':
        cfg = load_config()
        # 비밀번호/시크릿은 마스킹하여 반환
        safe_cfg = {
            'app_key': cfg.get('app_key', ''),
            'app_secret': '****' if cfg.get('app_secret') else '',
            'account': cfg.get('account', ''),
            'account_pw': '****' if cfg.get('account_pw') else '',
            'configured': bool(cfg.get('app_key') and cfg.get('app_secret') and
                               cfg.get('account') and cfg.get('account_pw'))
        }
        return jsonify(safe_cfg)

    # POST: 설정 저장
    data = request.get_json()
    cfg = load_config()  # 기존 설정 유지 (부분 업데이트 지원)

    for key in ('app_key', 'app_secret', 'account', 'account_pw'):
        val = data.get(key, '')
        # '****'가 들어오면 기존 값 유지 (마스킹된 값 그대로 보낸 경우)
        if val and val != '****':
            cfg[key] = val

    save_config(cfg)
    return jsonify({'success': True, 'message': 'LS증권 API 설정이 저장되었습니다.'})


@app.route('/api/ls/fetch-trades', methods=['POST'])
def ls_fetch_trades():
    """
    LS증권 API로 체결 내역 조회
    Body: { from_date: "YYYYMMDD", to_date: "YYYYMMDD", stock_code: "" }
    """
    try:
        from ls_api import fetch_trade_history
    except ImportError:
        return jsonify({'error': 'ls_api 모듈을 찾을 수 없습니다.'}), 500

    data = request.get_json() or {}
    from_date = data.get('from_date', '')
    to_date = data.get('to_date', '')
    stock_code = data.get('stock_code', '')

    if not from_date or not to_date:
        return jsonify({'error': '조회 기간(from_date, to_date)을 입력하세요. 형식: YYYYMMDD'}), 400

    try:
        trades = fetch_trade_history(
            from_date=from_date,
            to_date=to_date,
            stock_code=stock_code
        )
        return jsonify({
            'success': True,
            'count': len(trades),
            'from_date': from_date,
            'to_date': to_date,
            'trades': trades
        })
    except ValueError as e:
        # 설정 미완료 등 사용자 입력 오류
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': f'LS API 조회 실패: {str(e)}'}), 500


@app.route('/api/ls/import-trades', methods=['POST'])
def ls_import_trades():
    """
    선택한 거래내역을 Excel DB(매매일지 시트)에 저장
    Body: {
        file: "파일경로",
        trades: [{ date, name, qty, price, type, investment, memo }, ...]
    }
    """
    if not ONEDRIVE_PATH:
        return jsonify({'error': 'OneDrive 경로가 설정되지 않았습니다.'}), 400

    data = request.get_json() or {}
    file_path = data.get('file', '')
    trades = data.get('trades', [])

    if not file_path:
        return jsonify({'error': 'file 경로를 지정하세요.'}), 400
    if not trades:
        return jsonify({'error': '저장할 거래 내역이 없습니다.'}), 400

    full_path = os.path.join(ONEDRIVE_PATH, file_path)
    if not os.path.isfile(full_path):
        return jsonify({'error': f'파일을 찾을 수 없습니다: {file_path}'}), 404

    try:
        wb = openpyxl.load_workbook(full_path, rich_text=True)
        try:
            sheet_name = '매매일지'
            if sheet_name not in wb.sheetnames:
                return jsonify({'error': f'"{sheet_name}" 시트를 찾을 수 없습니다.'}), 404

            ws = wb[sheet_name]
            saved_count = 0
            errors = []

            for trade in trades:
                try:
                    # 기존 매매일지 컬럼 순서: [날짜, 종목, 수량, 단가, 매매종류, 투자금]
                    row_values = [
                        trade.get('date', ''),
                        trade.get('name', ''),
                        int(trade.get('qty', 0)),
                        int(trade.get('price', 0)),
                        trade.get('type', '매수'),
                        float(trade.get('investment', 0)),
                    ]
                    # 메모가 있으면 7번째 컬럼에 추가
                    memo = trade.get('memo', '')
                    if memo:
                        row_values.append(memo)

                    ws.append(row_values)

                    # 폰트 색상 (매도=빨강, 매수=검정)
                    last_row = ws.max_row
                    font_color = "FF0000" if trade.get('type') == '매도' else "000000"
                    cell_font = Font(color=font_color)
                    for col_idx in range(1, len(row_values) + 1):
                        ws.cell(row=last_row, column=col_idx).font = cell_font

                    # 포트폴리오 맵 동기화
                    stock_name = trade.get('name', '')
                    investment = float(trade.get('investment', 0))
                    if stock_name:
                        try:
                            sync_portfolio_map(wb, stock_name, investment)
                        except Exception:
                            pass  # 맵 동기화 실패는 저장 자체를 막지 않음

                    saved_count += 1

                except Exception as row_e:
                    errors.append(f"{trade.get('name', '?')}: {str(row_e)}")

            wb.save(full_path)
        finally:
            try: wb.close()
            except: pass

        # 아이폰 앱용 JSON 자동 갱신
        trigger_export()

        msg = f'{saved_count}건 저장 완료'
        if errors:
            msg += f' (오류 {len(errors)}건: {"; ".join(errors[:3])})'

        return jsonify({'success': True, 'saved': saved_count, 'errors': errors, 'message': msg})

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': f'저장 오류: {str(e)}'}), 500

TARGET_PRICES_FILE = os.path.join("StockPortfolioApp", "public", "data", "target_prices.json")

@app.route('/api/target-prices', methods=['GET'])
def get_target_prices():
    try:
        if os.path.exists(TARGET_PRICES_FILE):
            with open(TARGET_PRICES_FILE, 'r', encoding='utf-8') as f:
                return jsonify(json.load(f))
        return jsonify({})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/target-prices', methods=['POST'])
def save_target_prices():
    try:
        data = request.get_json()
        os.makedirs(os.path.dirname(TARGET_PRICES_FILE), exist_ok=True)
        with open(TARGET_PRICES_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        # 파일이 업데이트되었으므로 GitHub 배포 스크립트 실행 (백그라운드)
        trigger_export()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/ls/current-prices', methods=['POST'])
def ls_current_prices():
    """
    LS증권 API로 여러 종목의 현재가 조회
    Body: { "shcodes": ["005930", ...], "names": ["삼성전자", "SK하이닉스", ...] }
    반환: { "005930": 80000, "삼성전자": 80000 }
    """
    try:
        from ls_api import fetch_current_prices, get_stock_codes_by_names, load_config, get_access_token
    except ImportError:
        return jsonify({'error': 'ls_api 모듈을 찾을 수 없습니다.'}), 500

    data = request.get_json() or {}
    shcodes = data.get('shcodes', [])
    names = data.get('names', [])
    
    name_to_code = {}
    code_to_name = {}

    if names:
        cfg = load_config()
        token = get_access_token(cfg["app_key"], cfg["app_secret"])
        if token:
            name_to_code = get_stock_codes_by_names(token, names)
            shcodes.extend(name_to_code.values())
            code_to_name = {v: k for k, v in name_to_code.items()}

    shcodes = list(set(shcodes)) # 중복 제거

    if not shcodes:
        return jsonify({'error': '종목 코드(shcodes)나 종목명(names)을 제공해주세요.'}), 400

    try:
        prices = fetch_current_prices(shcodes)
        
        # 이름으로 요청된 경우 이름으로도 가격 추가
        result_prices = {**prices}
        for code, price in prices.items():
            if code in code_to_name:
                result_prices[code_to_name[code]] = price
                
        return jsonify({
            'success': True,
            'prices': result_prices
        })
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': f'현재가 조회 실패: {str(e)}'}), 500


@app.route('/api/ls/moving-averages', methods=['GET'])
def ls_moving_averages():
    """
    단일 종목의 이동평균선(5, 20, 60, 120일) 데이터 조회
    Query: ?shcode=005930 또는 ?name=삼성전자
    """
    try:
        from ls_api import fetch_moving_averages, get_stock_codes_by_names, load_config, get_access_token
    except ImportError:
        return jsonify({'error': 'ls_api 모듈을 찾을 수 없습니다.'}), 500

    shcode = request.args.get('shcode', '')
    name = request.args.get('name', '')

    if name and not shcode:
        cfg = load_config()
        token = get_access_token(cfg["app_key"], cfg["app_secret"])
        if token:
            name_to_code = get_stock_codes_by_names(token, [name])
            if name in name_to_code:
                shcode = name_to_code[name]

    if not shcode:
        return jsonify({'error': '종목 코드나 이름을 제공해주세요.'}), 400

    try:
        ma_data = fetch_moving_averages(shcode)
        return jsonify({
            'success': True,
            'data': ma_data
        })
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': f'이동평균선 조회 실패: {str(e)}'}), 500


# ── 신호 데이터 원격 갱신 (모바일 → PC 서버 트리거) ──
_signal_refresh_lock = threading.Lock()

@app.route('/api/refresh-signals', methods=['GET', 'POST'])
def refresh_signals():
    """
    모바일에서 호출하여 신호 데이터(이평선/RSI)를 갱신하고 GitHub에 반영.
    CORS 프록시 없이 모바일에서도 최신 시세를 확인할 수 있게 합니다.
    
    동작 흐름:
    1. export_signals.py 실행 (LS증권 API로 최신 이평선/RSI 수집)
    2. StockPortfolioApp/public/data → mobile/data 복사
    3. Git add + commit + push (GitHub Pages 반영)
    4. 완료 시 JSON 응답 반환
    """
    if not _signal_refresh_lock.acquire(blocking=False):
        return jsonify({
            'success': False,
            'message': '이미 신호 데이터 갱신이 진행 중입니다. 잠시 후 다시 시도하세요.'
        }), 429  # Too Many Requests

    try:
        import time as _time

        # 1단계: export_signals.py 실행 (이평선/RSI 수집)
        print("[REFRESH-SIGNALS] 1. 신호 데이터(이평선/RSI) 수집 시작...")
        result_sig = subprocess.run(
            [sys.executable, os.path.join(BASE_DIR, "export_signals.py")],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
            cwd=BASE_DIR, timeout=600  # 종목 수에 따라 시간이 걸릴 수 있음
        )
        if result_sig.returncode != 0:
            print(f"[REFRESH-SIGNALS] ❌ 수집 실패: {result_sig.stderr[:500]}")
            return jsonify({
                'success': False,
                'message': f'신호 데이터 수집 실패: {result_sig.stderr[:200]}'
            }), 500
        print("[REFRESH-SIGNALS] ✅ 신호 데이터 수집 완료")

        # 2단계: mobile/data 로 복사
        print("[REFRESH-SIGNALS] 2. 모바일 데이터 복사 중...")
        src = os.path.join(BASE_DIR, "StockPortfolioApp", "public", "data")
        dst = os.path.join(BASE_DIR, "mobile", "data")
        if os.path.exists(src):
            subprocess.run(
                f'xcopy "{src}" "{dst}" /E /I /Y',
                shell=True, capture_output=True, timeout=30
            )
        print("[REFRESH-SIGNALS] ✅ 모바일 데이터 복사 완료")

        # 3단계: Git push
        print("[REFRESH-SIGNALS] 3. Git push 시작...")
        subprocess.run(["git", "add", "."], cwd=BASE_DIR, capture_output=True, timeout=30)
        
        # 변경 사항이 없는 경우 커밋 건너뜀
        if not git_has_changes():
            print("[REFRESH-SIGNALS] 변경사항 없음 (이미 최신)")
            return jsonify({
                'success': True,
                'message': '신호 데이터가 이미 최신 상태입니다.',
                'pushed': False
            })

        from datetime import datetime as dt
        commit_msg = f"Signal refresh {dt.now().strftime('%Y-%m-%d %H:%M')}"
        commit_result = subprocess.run(
            ["git", "commit", "-m", commit_msg],
            cwd=BASE_DIR, capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=30
        )
        if commit_result.returncode != 0:
            print(f"[REFRESH-SIGNALS] ❌ Git 커밋 실패: {commit_result.stderr[:300]}")
            return jsonify({
                'success': False,
                'message': '신호 데이터 커밋 실패'
            }), 500

        env = os.environ.copy()
        env["GCM_INTERACTIVE"] = "never"
        env["GIT_TERMINAL_PROMPT"] = "0"
        
        # push 직전 rebase pull 실행
        print("[REFRESH-SIGNALS] push 전 원격 변경 사항 병합 (rebase)...")
        subprocess.run(
            ["git", "pull", "--rebase", "origin", "main"],
            cwd=BASE_DIR, capture_output=True, env=env, timeout=60
        )

        push_result = subprocess.run(
            ["git", "push", "origin", "main"],
            cwd=BASE_DIR, capture_output=True, text=True, encoding="utf-8", errors="replace",
            timeout=120, env=env
        )
        
        if push_result.returncode == 0:
            print("[REFRESH-SIGNALS] ✅ GitHub push 완료!")
            return jsonify({
                'success': True,
                'message': '신호 데이터 갱신 및 GitHub 반영 완료! 약 10초 후 새로고침하면 최신 데이터가 표시됩니다.',
                'pushed': True
            })
        else:
            print(f"[REFRESH-SIGNALS] ❌ GitHub push 실패: {push_result.stderr[:300]}")
            return jsonify({
                'success': True,
                'message': '신호 데이터 수집은 완료했으나, GitHub 반영에 실패했습니다. 잠시 후 자동 업로더가 처리합니다.',
                'pushed': False
            })

    except subprocess.TimeoutExpired:
        print("[REFRESH-SIGNALS] ❌ 시간 초과")
        return jsonify({
            'success': False,
            'message': '신호 데이터 수집 시간 초과 (10분). 종목 수가 너무 많을 수 있습니다.'
        }), 500
    except Exception as e:
        print(f"[REFRESH-SIGNALS] ❌ 예외 발생: {e}")
        return jsonify({
            'success': False,
            'message': f'오류 발생: {str(e)}'
        }), 500
    finally:
        if _signal_refresh_lock.locked():
            _signal_refresh_lock.release()


if __name__ == '__main__':
    print("=" * 60)
    print("  Stock Portfolio Analysis Server")
    print("=" * 60)
    if ONEDRIVE_PATH:
        print(f"  OneDrive Path: {os.path.normpath(ONEDRIVE_PATH)}")
    else:
        print("  OneDrive path not detected.")
        print("  Please set the path in the Web UI.")
    print(f"  Server Address: http://localhost:5000")
    print("=" * 60)
    
    try:
        # 배포 시에는 debug=False 권장
        app.run(debug=False, port=5000, host='0.0.0.0')
    except Exception as e:
        print(f"Error starting server: {e}")
        sys.exit(1)
