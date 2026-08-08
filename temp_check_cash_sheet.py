import openpyxl
import os

excel_path = r"C:\Users\zerod\OneDrive\주식 체크 리스트_20220328.xlsx"
print(f"파일 존재: {os.path.exists(excel_path)}")

wb = openpyxl.load_workbook(excel_path)
print('시트 목록:', wb.sheetnames)

sh = '현금비중'
print(f'"{sh}" 존재:', sh in wb.sheetnames)

if sh in wb.sheetnames:
    ws = wb[sh]
    print(f'행 수: {ws.max_row}, 열 수: {ws.max_column}')
    for r in range(1, min(ws.max_row+1, 10)):
        row_data = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
        print(f'  Row {r}: {row_data}')
else:
    print(f'"{sh}" 시트가 없습니다. 새로 생성 필요.')
