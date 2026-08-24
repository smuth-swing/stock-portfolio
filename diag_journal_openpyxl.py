# -*- coding: utf-8 -*-
"""매매일지 시트 openpyxl 원본 구조 재진단"""
import io, sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
FULL = r"C:\Users\zerod\OneDrive\주식 체크 리스트_20220328.xlsx"
with open(FULL, 'rb') as f:
    file_data = f.read()

wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True, rich_text=True)
ws = wb['매매일지']
print("max_row=%d max_col=%d" % (ws.max_row, ws.max_column))
print("merged:", list(ws.merged_cells.ranges)[:10])
for r in range(1, min(5, ws.max_row) + 1):
    vals = [repr(ws.cell(row=r, column=c).value)[:30] for c in range(1, 12)]
    print(f"Row {r}: {vals}")
print("...마지막 3행:")
for r in range(max(1, ws.max_row - 2), ws.max_row + 1):
    vals = [repr(ws.cell(row=r, column=c).value)[:30] for c in range(1, 12)]
    print(f"Row {r}: {vals}")
