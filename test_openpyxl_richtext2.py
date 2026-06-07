import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

wb = openpyxl.Workbook()
ws = wb.active
rt = CellRichText(
    TextBlock(InlineFont(rFont='맑은 고딕'), 'Hello '),
    TextBlock(InlineFont(rFont='맑은 고딕', strike=True), 'world')
)
ws['A1'].value = rt
wb.save('test_rich_text2.xlsx')

wb2 = openpyxl.load_workbook('test_rich_text2.xlsx')
cell = wb2.active['A1']
print("Type:", type(cell.value))
if isinstance(cell.value, CellRichText):
    for elem in cell.value:
        print("Text:", elem.text, "Strike:", getattr(elem.font, 'strike', False))
else:
    print("Value:", cell.value)
