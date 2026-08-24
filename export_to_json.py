"""
export_to_json.py — 엑셀 데이터를 JSON으로 내보내기 스크립트
- server.py와 동일하게 로컬 OneDrive 폴더에서 파일을 직접 읽음
- 별도 인증(Graph API 토큰) 불필요
- 4개 시트를 JSON으로 변환하여 OneDrive 공유 폴더에 저장
- Windows 작업 스케줄러로 하루 1회 자동 실행 가능
"""

import io
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.cell.rich_text import CellRichText
import socket

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"

# Windows 터미널 UTF-8 출력 설정
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

# ==================== 설정 ====================
TARGET_FILE = '주식 체크 리스트_20220328.xlsx'

# OneDrive 로컬 동기화 경로 (server.py와 동일)
ONEDRIVE_PATH = r'C:\Users\zerod\OneDrive'

# JSON 저장 폴더 (OneDrive 내부 → 자동 클라우드 동기화됨)
OUTPUT_DIR = Path(ONEDRIVE_PATH) / '주식앱데이터'

# 아이폰 앱(Expo) 로컬 서버용 폴더 (앱 내부 데이터 제공용)
APP_DATA_DIR = Path(__file__).parent / 'StockPortfolioApp' / 'public' / 'data'

# 내보낼 시트 목록 (시트명 → JSON 파일명)
EXPORT_SHEETS = {
    '매매일지':     'trade_journal.json',
    '포트폴리오 맵': 'portfolio_map.json',
    '탐구생활':     'investigation.json',
    '실적':         'performance.json',
    '현금비중':     'cash_snapshots.json',
}

# 추가 내보내기: 신호 데이터 (PC에서 계산한 목표가 크로스 상태)
# 모바일 앱이 PC와 동일한 신호를 표시하기 위해 사용
def export_investigation_signals(file_data, sheet_names):
    """탐구생활 시트의 목표일/목표가 기반 신호 데이터 생성"""
    import re
    from datetime import date
    
    if '탐구생활' not in sheet_names:
        return {}
    
    today_str = date.today().isoformat()
    signals = {}
    
    try:
        df = pd.read_excel(io.BytesIO(file_data), sheet_name='탐구생활', engine='openpyxl')
        df = df.fillna('')
        
        cols = [str(c) for c in df.columns.tolist()]
        name_col = next((c for c in cols if '종목명' in c or c == 'Unnamed: 1'), None)
        td_col = next((c for c in cols if '목표일' in c or c == 'Unnamed: 8'), None)
        tp_col = next((c for c in cols if '목표가' in c or c == 'Unnamed: 9' or c == 'Unnamed: 10'), None)
        
        if not name_col:
            return {}
        
        for _, row in df.iterrows():
            stock_name = str(row.get(name_col, '')).replace('~~', '').strip()
            if not stock_name or stock_name in ('종목', 'stock'):
                continue
            
            target_date = str(row.get(td_col, '')).strip() if td_col else ''
            target_price_raw = str(row.get(tp_col, '')).strip() if tp_col else ''
            target_price = int(re.sub(r'[^0-9]', '', target_price_raw)) if target_price_raw else 0
            
            has_date_signal = False
            date_match = re.search(r'(\d{4}-\d{2}-\d{2})', target_date)
            if date_match and date_match.group(1) <= today_str:
                has_date_signal = True
            
            # 목표가 신호: 모바일은 실시간 가격 조회 불가 → PC에서 _priceCrossTracker 기반 판단
            # 여기서는 일단 목표가 존재 여부만 기록 (실제 크로스 여부는 PC 웹앱에서만 판단 가능)
            has_price_signal = False  # 모바일 단독으로는 판단 불가 → PC 연동 필요
            
            signals[stock_name] = {
                'hasDateSignal': has_date_signal,
                'hasPriceSignal': has_price_signal,
                'targetDate': target_date,
                'targetPrice': target_price,
            }
    except Exception as e:
        print(f'  [경고] 신호 데이터 생성 실패: {e}')
    
    return signals

def export_cash_accounts(file_data, sheet_names):
    """Excel '_계좌정보' 시트에서 현금 계좌 목록 추출"""
    if '_계좌정보' not in sheet_names:
        return []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
        ws = wb['_계좌정보']
        data = []
        in_section = False
        for r in range(1, ws.max_row + 1):
            cell_val = str(ws.cell(r, 1).value or '').strip()
            if cell_val == '현금계좌':
                in_section = True
                continue
            if in_section:
                if cell_val == '' or cell_val.startswith('#'):
                    break
                name = str(ws.cell(r, 1).value or '').strip()
                amount = ws.cell(r, 2).value or 0
                if name:
                    data.append({'name': name, 'amount': float(amount)})
        wb.close()
        return data
    except Exception as e:
        print(f'  [경고] 현금 계좌 데이터 추출 실패: {e}')
        return []

# ==================== 취소선 텍스트 추출 (server.py와 동일) ====================
def extract_rich_text(cell):
    """엑셀 셀의 취소선 서식을 ~~텍스트~~ 마크다운으로 변환"""
    if not hasattr(cell, 'value') or cell.value is None:
        return ''
    if isinstance(cell.value, CellRichText):
        result = []
        for part in cell.value:
            text = part.text if hasattr(part, 'text') else str(part)
            if hasattr(part, 'font') and part.font and part.font.strike:
                result.append(f'~~{text}~~')
            else:
                result.append(text)
        return ''.join(result)
    if cell.font and cell.font.strike:
        return f'~~{cell.value}~~'
    return str(cell.value)


# ==================== 시트 → JSON 변환 (server.py read_excel 로직 재현) ====================
def sheet_to_json(file_data: bytes, sheet_name: str, sheet_names: list) -> dict:
    """
    server.py의 read_excel 엔드포인트와 동일한 로직으로
    엑셀 시트를 JSON 구조로 변환합니다.
    """
    df = pd.read_excel(io.BytesIO(file_data), sheet_name=sheet_name, engine='openpyxl')
    df = df.fillna('')

    # --- 헤더 자동 탐색 및 승격 (첫 행이 빈 줄이거나 Unnamed인 경우) ---
    header_row_idx = 0
    is_unnamed_header = any(str(c).startswith('Unnamed:') for c in df.columns)
    if not df.empty and (is_unnamed_header or '실적' in sheet_name or '매매' in sheet_name):
        keywords = ['Date', '종목', '날짜', '수량', '가격', '매매유형', '연도', '수익율', '종목명']
        for i in range(min(10, len(df))):
            row_vals = [str(x).strip() for x in df.iloc[i].values]
            if any(any(k in val for k in keywords) for val in row_vals if val):
                header_row_idx = i
                new_cols = []
                for j, val in enumerate(row_vals):
                    if val and val != 'nan':
                        new_cols.append(val)
                    else:
                        new_cols.append(f'Unnamed: {j}')
                df.columns = new_cols
                df = df.iloc[i + 1:].reset_index(drop=True)
                break

    # 숫자형 컬럼 강제 변환
    for col in df.columns:
        if '연도' in str(col):
            continue
        try:
            temp_numeric = pd.to_numeric(df[col].replace('', pd.NA), errors='coerce')
            if temp_numeric.notna().any():
                if temp_numeric.notna().sum() / len(df) > 0.2:
                    df[col] = temp_numeric.fillna(0)
        except Exception:
            pass

    numeric_columns = df.select_dtypes(include=['number']).columns.tolist()
    columns = [str(c) for c in df.columns.tolist()]

    # 탐구생활 / 실적 시트: 취소선 서식 보존
    is_special = any(k in sheet_name for k in ['탐구', '실적'])
    if is_special:
        wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True, rich_text=True)
        ws = wb[sheet_name]
        start_row = header_row_idx + 3 if '실적' in sheet_name else 2
        data = []
        for r_idx in range(start_row, ws.max_row + 1):
            row_data = {'_realIndex': r_idx - 2}
            for c_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                row_data[col_name] = extract_rich_text(cell)
            data.append(row_data)
    else:
        data = []
        for idx, row in df.iterrows():
            row_data = {'_realIndex': idx}
            for i, col in enumerate(df.columns):
                val = row[col]
                col_name = columns[i]
                if pd.isna(val):
                    row_data[col_name] = ''
                elif isinstance(val, (int, float)):
                    row_data[col_name] = val
                else:
                    row_data[col_name] = str(val)
            data.append(row_data)

    return {
        'file_name': TARGET_FILE,
        'sheet_names': sheet_names,
        'current_sheet': sheet_name,
        'columns': columns,
        'numeric_columns': [str(c) for c in numeric_columns],
        'data': data,
        'row_count': len(data),
    }


# ==================== 메인 내보내기 함수 ====================
def export_all():
    print('=' * 60)
    print('  주식 포트폴리오 데이터 JSON 내보내기')
    print('=' * 60)

    # 1. 엑셀 파일 경로 확인
    excel_path = Path(ONEDRIVE_PATH) / TARGET_FILE
    if not excel_path.is_file():
        print(f'[오류] 파일을 찾을 수 없습니다: {excel_path}')
        print('  OneDrive 동기화가 완료되었는지 확인하세요.')
        return False

    print(f'[OK] 파일 확인: {excel_path}')

    # 2. 파일을 메모리로 읽기 (잠금 방지)
    with open(excel_path, 'rb') as f:
        file_data = f.read()

    # 시트 목록 가져오기
    xl = pd.ExcelFile(io.BytesIO(file_data), engine='openpyxl')
    sheet_names = xl.sheet_names
    print(f'[OK] 시트 목록: {sheet_names}')

    # 3. 출력 폴더 생성
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    APP_DATA_DIR.mkdir(parents=True, exist_ok=True)
    print(f'[OK] 원드라이브 저장 폴더: {OUTPUT_DIR}')
    print(f'[OK] 앱 로컬 저장 폴더: {APP_DATA_DIR}')
    print('')

    # 4. 각 시트를 JSON으로 변환 및 저장
    success_count = 0
    for sheet_name, filename in EXPORT_SHEETS.items():
        if sheet_name not in sheet_names:
            print(f'  [건너뜀] 시트 없음: {sheet_name}')
            continue
        try:
            print(f'  변환 중: {sheet_name} ...', end=' ')
            
            # 현금비중 시트는 단순 배열 형식으로 내보내기 (모바일 차트용)
            if sheet_name == '현금비중':
                json_data = []
                wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
                ws = wb[sheet_name]
                for r in range(2, ws.max_row + 1):
                    month = ws.cell(r, 1).value
                    if not month:
                        continue
                    json_data.append({
                        'month': str(month).strip(),
                        'investment': float(ws.cell(r, 2).value or 0),
                        'cash': float(ws.cell(r, 3).value or 0),
                        'totalAsset': float(ws.cell(r, 4).value or 0),
                        'ratio': float(ws.cell(r, 5).value or 0)
                    })
                wb.close()
                json_data.sort(key=lambda x: x['month'])
                row_count = len(json_data)
                is_simple_array = True
            else:
                json_data = sheet_to_json(file_data, sheet_name, sheet_names)
                row_count = json_data.get("row_count", 0)
                is_simple_array = False
            
            output_path = OUTPUT_DIR / filename
            app_path = APP_DATA_DIR / filename
            
            # OneDrive 폴더에 저장
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, ensure_ascii=False, separators=(',', ':'), default=str)
                
            # 앱 public 폴더에도 복사 저장
            with open(app_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, ensure_ascii=False, separators=(',', ':'), default=str)
                
            size_kb = output_path.stat().st_size / 1024
            print(f'완료! ({row_count}행, {size_kb:.1f} KB) -> {filename}')
            success_count += 1
        except Exception as e:
            print(f'실패!')
            print(f'     오류: {e}')

    # 4.5. 현금계좌 및 탐구생활 신호 데이터 내보내기 (모바일용)
    if '_계좌정보' in sheet_names:
        try:
            print(f'  현금 계좌 데이터 생성 중...', end=' ')
            cash_accs = export_cash_accounts(file_data, sheet_names)
            acc_path = APP_DATA_DIR / 'cash_accounts.json'
            with open(acc_path, 'w', encoding='utf-8') as f:
                json.dump(cash_accs, f, ensure_ascii=False, separators=(',', ':'), default=str)
            # OneDrive에도 복사
            acc_out = OUTPUT_DIR / 'cash_accounts.json'
            with open(acc_out, 'w', encoding='utf-8') as f:
                json.dump(acc_out_data if 'acc_out_data' in locals() else cash_accs, f, ensure_ascii=False, separators=(',', ':'), default=str)
            print(f'완료! ({len(cash_accs)}개 계좌) -> cash_accounts.json')
        except Exception as e:
            print(f'실패! 오류: {e}')

    if '탐구생활' in sheet_names:
        try:
            print(f'  신호 데이터 생성 중...', end=' ')
            sig_data = export_investigation_signals(file_data, sheet_names)
            sig_path = APP_DATA_DIR / 'investigation_signals.json'
            with open(sig_path, 'w', encoding='utf-8') as f:
                json.dump(sig_data, f, ensure_ascii=False, separators=(',', ':'), default=str)
            # OneDrive에도 복사
            sig_out = OUTPUT_DIR / 'investigation_signals.json'
            with open(sig_out, 'w', encoding='utf-8') as f:
                json.dump(sig_data, f, ensure_ascii=False, separators=(',', ':'), default=str)
            print(f'완료! ({len(sig_data)}종목) -> investigation_signals.json')
        except Exception as e:
            print(f'실패!')
            print(f'     오류: {e}')

    # 5. 메타데이터 저장 (앱에서 마지막 업데이트 시간 표시용)
    meta = {
        'updated_at': datetime.now(timezone.utc).isoformat(timespec='milliseconds').replace('+00:00', 'Z'),
        'updated_at_display': datetime.now().strftime('%Y년 %m월 %d일 %H:%M'),
        'exported_sheets': list(EXPORT_SHEETS.keys()),
        'success_count': success_count,
        'total_count': len(EXPORT_SHEETS),
        'source_file': TARGET_FILE,
        'server_ip': get_local_ip(),
    }
    with open(OUTPUT_DIR / 'meta.json', 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    with open(APP_DATA_DIR / 'meta.json', 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

    print('')
    print(f'[완료] {success_count}/{len(EXPORT_SHEETS)}개 시트 내보내기 성공')
    print(f'  업데이트 시각: {meta["updated_at_display"]}')
    print(f'  저장 위치: {OUTPUT_DIR}')
    print('')
    print('OneDrive 동기화 후 아이폰 앱에서 새로고침하면 최신 데이터를 확인할 수 있습니다.')
    print('=' * 60)
    return success_count > 0


if __name__ == '__main__':
    result = export_all()
    sys.exit(0 if result else 1)
