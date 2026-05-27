"""
OneDrive 엑셀 데이터 분석 서버 - Microsoft Graph API 버전
- Microsoft Graph API를 사용하여 클라우드의 엑셀 파일에 직접 접근
- OAuth 2.0 인증으로 안전한 클라우드 접근
- 로컬 동기화 지연 문제 완전 해결
- Flask 기반 REST API 서버
"""

import os
import sys
import json
import time
import io
import logging
import re
from pathlib import Path
from flask import Flask, jsonify, request, send_from_directory, redirect, url_for
from flask_cors import CORS
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.cell.rich_text import CellRichText, TextBlock
from datetime import datetime, timedelta
import requests
from dotenv import load_dotenv

# .env 파일 로드
load_dotenv()

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('server_graph.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# ==================== 설정 클래스 ====================
class Config:
    def __init__(self):
        self.TENANT_ID = os.getenv('TENANT_ID', 'common')
        self.CLIENT_ID = os.getenv('CLIENT_ID', '')
        self.CLIENT_SECRET = os.getenv('CLIENT_SECRET', '')
        self.REDIRECT_URI = os.getenv('REDIRECT_URI', 'http://localhost:5000/auth/callback')
        self.TOKEN_FILE = Path('.') / 'token.json'
        self.EXCEL_FILE_PATH = '주식 체크 리스트_20220328.xlsx'

config = Config()

# ==================== 유틸리티 함수 ====================
def parse_strikethrough_text(text):
    """
    텍스트 내의 <del>태그 또는 ~~패턴을 파싱하여 CellRichText 객체로 변환.
    한글 깨짐 방지를 위해 기본 폰트 설정을 유지합니다.
    """
    if not isinstance(text, str) or not text:
        return text
    
    # <del>태그 또는 ~~패턴 처리
    # 먼저 <del>... </del> 형식을 ~~...~~ 형식으로 통일하여 처리
    text = re.sub(r'<del>(.*?)</del>', r'~~\1~~', text)
    
    pattern = r'~~(.*?)~~'
    parts = re.split(pattern, text)
    
    if len(parts) == 1:
        return text
    
    rich_text = CellRichText()
    # 한글 깨짐 방지를 위한 폰트 설정 (맑은 고딕 등 기본 폰트 명시)
    default_font = Font(name='맑은 고딕')
    strike_font = Font(name='맑은 고딕', strike=True)
    
    for i, part in enumerate(parts):
        if not part: continue
        if i % 2 == 0:
            rich_text.add(TextBlock(default_font, part))
        else:
            rich_text.add(TextBlock(strike_font, part))
    
    return rich_text

def extract_rich_text(cell):
    """
    엑셀 셀의 실제 취소선 서식을 감지하여 마크다운(~~) 또는 HTML 형식이 포함된 문자열로 변환
    """
    if not hasattr(cell, 'value') or cell.value is None:
        return ""
    
    # RichText인 경우
    if isinstance(cell.value, CellRichText):
        result = []
        for part in cell.value:
            text = part.text if hasattr(part, 'text') else str(part)
            # 폰트 서식에 취소선이 있는지 확인
            if hasattr(part, 'font') and part.font and part.font.strike:
                result.append(f"~~{text}~~")
            else:
                result.append(text)
        return "".join(result)
    
    # 일반 텍스트인 경우 셀 전체 폰트 확인
    if cell.font and cell.font.strike:
        return f"~~{cell.value}~~"
        
    return str(cell.value)

# 글로벌 토큰
access_token = None
token_expiry = None


# ==================== 인증 관련 함수 ====================
def get_auth_url():
    """Azure AD 인증 URL 생성"""
    return (
        f"https://login.microsoftonline.com/{config.TENANT_ID}/oauth2/v2.0/authorize?"
        f"client_id={config.CLIENT_ID}"
        f"&redirect_uri={config.REDIRECT_URI}"
        f"&response_type=code"
        f"&scope=https://graph.microsoft.com/.default offline_access"
    )


def get_token_from_code(auth_code):
    """인증 코드로부터 액세스 토큰 획득"""
    token_url = f"https://login.microsoftonline.com/{config.TENANT_ID}/oauth2/v2.0/token"
    
    payload = {
        'client_id': config.CLIENT_ID,
        'client_secret': config.CLIENT_SECRET,
        'code': auth_code,
        'redirect_uri': config.REDIRECT_URI,
        'grant_type': 'authorization_code',
        'scope': 'https://graph.microsoft.com/.default'
    }
    
    try:
        response = requests.post(token_url, data=payload)
        if response.status_code == 200:
            token_data = response.json()
            save_token(token_data)
            return token_data.get('access_token')
        else:
            logger.error(f"토큰 획득 실패: {response.text}")
            return None
    except requests.RequestException as e:
        logger.error(f"토큰 요청 중 네트워크 오류: {e}")
        return None


def refresh_access_token():
    """리프레시 토큰으로 새 액세스 토큰 획득"""
    global access_token, token_expiry
    
    token_data = load_token()
    if not token_data or 'refresh_token' not in token_data:
        return False
    
    token_url = f"https://login.microsoftonline.com/{config.TENANT_ID}/oauth2/v2.0/token"
    
    payload = {
        'client_id': config.CLIENT_ID,
        'client_secret': config.CLIENT_SECRET,
        'refresh_token': token_data['refresh_token'],
        'grant_type': 'refresh_token',
        'scope': 'https://graph.microsoft.com/.default'
    }
    
    try:
        response = requests.post(token_url, data=payload)
        if response.status_code == 200:
            new_token_data = response.json()
            save_token(new_token_data)
            access_token = new_token_data.get('access_token')
            token_expiry = datetime.now() + timedelta(seconds=new_token_data.get('expires_in', 3600))
            return True
        else:
            logger.error(f"토큰 갱신 실패: {response.text}")
            return False
    except requests.RequestException as e:
        logger.error(f"토큰 갱신 중 네트워크 오류: {e}")
        return False


def save_token(token_data):
    """토큰을 파일에 저장"""
    with open(config.TOKEN_FILE, 'w') as f:
        json.dump(token_data, f, indent=2)
    logger.info(f"토큰 저장됨: {config.TOKEN_FILE}")


def load_token():
    """파일에서 토큰 로드"""
    if config.TOKEN_FILE.exists():
        with open(config.TOKEN_FILE, 'r') as f:
            return json.load(f)
    return None


def ensure_valid_token():
    """유효한 토큰 확보"""
    global access_token, token_expiry
    
    # 토큰이 없으면
    if not access_token:
        token_data = load_token()
        if not token_data:
            return False
        access_token = token_data.get('access_token')
        expires_in = token_data.get('expires_in', 3600)
        token_expiry = datetime.now() + timedelta(seconds=expires_in)
    
    # 토큰 만료 시간 체크 (만료 5분 전에 갱신)
    if token_expiry and datetime.now() > token_expiry - timedelta(minutes=5):
        logger.info("토큰 갱신 필요...")
        return refresh_access_token()
    
    return True


def get_headers():
    """API 요청 헤더 생성"""
    if not ensure_valid_token():
        return None
    return {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }


# ==================== Microsoft Graph API 함수 ====================
def find_file_by_name(filename):
    """파일명으로 파일 ID 찾기"""
    headers = get_headers()
    if not headers:
        return None
    
    # 쿼리로 파일 검색
    search_url = f"https://graph.microsoft.com/v1.0/me/drive/root/children"
    params = {
        '$filter': f"name eq '{filename}'"
    }
    
    response = requests.get(search_url, headers=headers, params=params)
    if response.status_code == 200:
        items = response.json().get('value', [])
        if items:
            return items[0]['id']
    
    logger.warning(f"파일을 찾을 수 없음: {filename}")
    return None


def download_file(file_id):
    """파일 다운로드"""
    headers = get_headers()
    if not headers:
        return None
    
    download_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{file_id}/content"
    response = requests.get(download_url, headers=headers)
    
    if response.status_code == 200:
        return io.BytesIO(response.content)
    else:
        logger.error(f"파일 다운로드 실패: {response.status_code}")
        return None


def upload_file(file_id, file_content):
    """파일 업로드 (덮어쓰기)"""
    headers = get_headers()
    if not headers:
        return False
    
    upload_url = f"https://graph.microsoft.com/v1.0/me/drive/items/{file_id}/content"
    
    response = requests.put(upload_url, headers=headers, data=file_content)
    
    if response.status_code in [200, 201]:
        print(f"파일 업로드 성공")
        return True
    else:
        print(f"파일 업로드 실패: {response.status_code} - {response.text}")
        return False


# ==================== Flask 라우트 ====================
@app.route('/')
def index():
    """메인 페이지 서빙"""
    return send_from_directory('.', 'index.html')


@app.route('/auth/login')
def login():
    """로그인 페이지로 리디렉트"""
    auth_url = get_auth_url()
    return redirect(auth_url)


@app.route('/auth/callback')
def auth_callback():
    """인증 콜백 처리"""
    auth_code = request.args.get('code')
    
    if not auth_code:
        return jsonify({'error': '인증 코드를 받지 못했습니다.'}), 400
    
    token = get_token_from_code(auth_code)
    if token:
        global access_token
        access_token = token
        return '''
        <html>
            <head><title>인증 성공</title></head>
            <body>
                <h1>✅ 인증 성공!</h1>
                <p>이제 애플리케이션을 사용할 수 있습니다.</p>
                <a href="/">돌아가기</a>
                <script>
                    setTimeout(() => {
                        window.location.href = '/';
                    }, 2000);
                </script>
            </body>
        </html>
        '''
    else:
        return jsonify({'error': '토큰 획득 실패'}), 500


@app.route('/api/onedrive-status')
def onedrive_status():
    """OneDrive 연결 상태 확인 (Graph API)"""
    headers = get_headers()
    
    if headers:
        # 사용자 정보 조회로 연결 확인
        response = requests.get('https://graph.microsoft.com/v1.0/me', headers=headers)
        if response.status_code == 200:
            user_info = response.json()
            return jsonify({
                'connected': True,
                'user': user_info.get('displayName', '사용자'),
                'email': user_info.get('mail', ''),
                'message': f"Microsoft Graph 연결됨: {user_info.get('displayName', '사용자')}"
            })
    
    return jsonify({
        'connected': False,
        'message': '인증 필요. /auth/login으로 이동하세요.'
    })


@app.route('/api/read-excel')
def read_excel():
    """엑셀 파일 데이터 읽기 (Graph API)"""
    file_path = request.args.get('file', EXCEL_FILE_PATH)
    sheet_name = request.args.get('sheet', None)
    
    # 파일 찾기
    file_id = find_file_by_name(file_path.split('/')[-1])  # 파일명만 추출
    if not file_id:
        return jsonify({'error': f'파일을 찾을 수 없습니다: {file_path}'}), 404
    
    # 파일 다운로드
    file_content = download_file(file_id)
    if not file_content:
        return jsonify({'error': '파일 다운로드 실패'}), 500
    
    try:
        # 엑셀 파일 읽기
        xl = pd.ExcelFile(file_content, engine='openpyxl')
        sheet_names = xl.sheet_names
        
        # 특정 시트 또는 첫 번째 시트 읽기
        target_sheet = sheet_name if sheet_name else sheet_names[0]
        df = pd.read_excel(file_content, sheet_name=target_sheet, engine='openpyxl')
        
        # NaN 값 처리
        df = df.fillna('')
        
        # 숫자형 컬럼 감지
        numeric_columns = df.select_dtypes(include=['number']).columns.tolist()
        
        # 기본 통계 계산
        stats = {}
        for col in numeric_columns:
            col_data = df[col].replace('', pd.NA).dropna()
            if len(col_data) > 0:
                stats[col] = {
                    'mean': round(float(col_data.mean()), 2),
                    'sum': round(float(col_data.sum()), 2),
                    'min': round(float(col_data.min()), 2),
                    'max': round(float(col_data.max()), 2),
                    'count': int(col_data.count()),
                    'std': round(float(col_data.std()), 2) if len(col_data) > 1 else 0
                }
        
        # 데이터를 JSON 직렬화 가능한 형태로 변환
        columns = [str(c) for c in df.columns.tolist()]
        
        # 탐구생활 시트인 경우 서식 데이터(취소선 등)를 위해 openpyxl로 재확인
        is_exploration = "탐구" in target_sheet
        if is_exploration:
            # openpyxl 워크북 로드
            file_content.seek(0)
            wb_format = openpyxl.load_workbook(file_content, data_only=False)
            ws_format = wb_format[target_sheet]
            
            data = []
            # 데이터 시작 행 (헤더 다음 행부터)
            for r_idx in range(2, ws_format.max_row + 1):
                row_data = {}
                for c_idx, col_name in enumerate(columns, start=1):
                    cell = ws_format.cell(row=r_idx, column=c_idx)
                    # 실제 취소선 서식 추출
                    row_data[col_name] = extract_rich_text(cell)
                data.append(row_data)
        else:
            # 일반 시트는 pandas 데이터 사용 (속도 우선)
            data = []
            for _, row in df.iterrows():
                row_data = {}
                for i, col in enumerate(df.columns):
                    val = row[col]
                    col_name = columns[i]
                    if pd.isna(val):
                        row_data[col_name] = ''
                    elif isinstance(val, (int, float)):
                        row_data[col_name] = val
                    else:
                        row_data[col_name] = str(val)
                data.append(row_data)

        return jsonify({
            'file_name': file_path.split('/')[-1],
            'last_modified': time.time(),
            'sheet_names': sheet_names,
            'current_sheet': target_sheet,
            'columns': columns,
            'numeric_columns': [str(c) for c in numeric_columns],
            'data': data,
            'row_count': len(data),
            'stats': stats
        })
    
    except Exception as e:
        return jsonify({'error': f'엑셀 파일 읽기 오류: {str(e)}'}), 500


@app.route('/api/save-journal', methods=['POST'])
def save_journal():
    """매매일지 데이터 추가 저장 (Graph API)"""
    data = request.get_json()
    file_path = data.get('file', EXCEL_FILE_PATH)
    sheet_name = data.get('sheet', '매매일지')
    row_values = data.get('row', [])  # [날짜, 종목, 수량, 단가, 매매종류, 투자금]
    
    try:
        # 파일 찾기
        file_id = find_file_by_name(file_path.split('/')[-1])
        if not file_id:
            return jsonify({'error': f'파일을 찾을 수 없습니다: {file_path}'}), 404
        
        # 파일 다운로드
        file_content = download_file(file_id)
        if not file_content:
            return jsonify({'error': '파일 다운로드 실패'}), 500
        
        # openpyxl로 엑셀 파일 열기
        wb = openpyxl.load_workbook(file_content)
        
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]
        
        # 데이터 추가
        ws.append(row_values)
        
        # 마지막 행의 폰트 색상 설정
        last_row = ws.max_row
        trade_type = row_values[4] if len(row_values) > 4 else ""
        
        # 색상 결정 (매도: 빨간색, 매수: 검정색)
        font_color = "FF0000" if trade_type == "매도" else "000000"
        cell_font = Font(color=font_color)
        
        for col_idx in range(1, len(row_values) + 1):
            ws.cell(row=last_row, column=col_idx).font = cell_font
        
        # 2. 포트폴리오 맵 동기화 로직
        try:
            trade_stock = str(row_values[1]).strip() if len(row_values) > 1 else ""
            trade_amount = float(row_values[5]) if len(row_values) > 5 else 0
            
            if trade_stock:
                sync_portfolio_map(wb, trade_stock, trade_amount)
                # 매칭된 실제 점 개수 기준으로 매매일지 금액 재확인
                final_ones = 0
                ws_map = wb['포트폴리오 맵']
                trade_stock_clean = trade_stock.replace(" ", "")
                for r in range(1, ws_map.max_row + 1):
                    cell_val = str(ws_map.cell(row=r, column=4).value or "").strip().replace(" ", "")
                    if cell_val == trade_stock_clean:
                        for c in range(5, 101):
                            if ws_map.cell(row=r, column=c).value == 1:
                                final_ones += 1
                        ws.cell(row=last_row, column=6).value = final_ones * 100
                        break
        except Exception as inner_e:
            print(f"Portfolio Map update error: {inner_e}")

        # 메모리에 파일 저장 및 업로드
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_bytes = output.getvalue()
        
        # Graph API로 파일 업로드 (덮어쓰기)
        if upload_file(file_id, file_bytes):
            return jsonify({'success': True, 'message': '클라우드에 성공적으로 저장되었습니다.'})
        else:
            return jsonify({'error': '파일 업로드 실패'}), 500
    
    except Exception as e:
        return jsonify({'error': f'저장 오류: {str(e)}'}), 500


@app.route('/api/update-row', methods=['POST'])
def update_row():
    data = request.get_json()
    file_path = data.get('file', EXCEL_FILE_PATH)
    sheet_name = data.get('sheet')
    row_index = int(data.get('rowIndex', 0))
    values = data.get('values', [])

    if sheet_name is None:
        return jsonify({'error': 'sheet 값이 필요합니다.'}), 400

    try:
        file_id = find_file_by_name(file_path.split('/')[-1])
        if not file_id:
            return jsonify({'error': f'파일을 찾을 수 없습니다: {file_path}'}), 404

        file_content = download_file(file_id)
        if not file_content:
            return jsonify({'error': '파일 다운로드 실패'}), 500

        wb = openpyxl.load_workbook(file_content)
        if sheet_name not in wb.sheetnames:
            return jsonify({'error': f'시트를 찾을 수 없습니다: {sheet_name}'}), 404

        ws = wb[sheet_name]
        target_row = row_index + 2
        for col_idx, value in enumerate(values, start=1):
            # 취소선 패턴 처리
            processed_value = parse_strikethrough_text(value)
            ws.cell(row=target_row, column=col_idx).value = processed_value

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        if upload_file(file_id, output.getvalue()):
            return jsonify({'success': True, 'message': '행이 업데이트되었습니다.'})
        return jsonify({'error': '파일 업로드 실패'}), 500
    except Exception as e:
        return jsonify({'error': f'업데이트 오류: {str(e)}'}), 500


def sync_portfolio_map(wb, stock_name, trade_amount):
    """포트폴리오 맵의 종목 마크(점)를 투자금액에 맞게 동기화"""
    if '포트폴리오 맵' not in wb.sheetnames:
        return
        
    ws_map = wb['포트폴리오 맵']
    stock_clean = stock_name.replace(" ", "")
    target_marks = int(trade_amount // 100)
    
    # 채우기 스타일 설정
    from openpyxl.styles import PatternFill
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    no_fill = PatternFill(fill_type=None)
    
    # 종목명이 일치하는 모든 행을 찾아서 처리
    for r in range(1, ws_map.max_row + 1):
        cell_val = str(ws_map.cell(row=r, column=4).value or "").strip().replace(" ", "")
        if cell_val and (stock_clean in cell_val or cell_val in stock_clean):
            # 1. 해당 행의 마크 영역(5~100컬럼) 초기화
            for c in range(5, 101):
                ws_map.cell(row=r, column=c).value = None
                ws_map.cell(row=r, column=c).fill = no_fill
            
            # 2. 투자금에 해당하는 개수만큼 점(1) 새로 찍기
            if target_marks > 0:
                for i in range(target_marks):
                    col_idx = 5 + i
                    if col_idx <= 100:
                        cell = ws_map.cell(row=r, column=col_idx)
                        cell.value = 1
                        cell.fill = yellow_fill
    print(f"📊 [{stock_name}] 포트폴리오 맵 동기화 완료: {target_marks}개 마크")


@app.route('/api/delete-row', methods=['POST'])
def delete_row():
    """행 삭제 API - 매매일지의 경우 이전 데이터로 동기화 포함"""
    data = request.get_json()
    file_path = data.get('file', EXCEL_FILE_PATH)
    sheet_name = data.get('sheet')
    row_index = int(data.get('rowIndex', 0))

    if sheet_name is None:
        return jsonify({'error': 'sheet 값이 필요합니다.'}), 400

    try:
        file_id = find_file_by_name(file_path.split('/')[-1])
        if not file_id:
            return jsonify({'error': f'파일을 찾을 수 없습니다: {file_path}'}), 404

        file_content = download_file(file_id)
        if not file_content:
            return jsonify({'error': '파일 다운로드 실패'}), 500

        wb = openpyxl.load_workbook(file_content)
        if sheet_name not in wb.sheetnames:
            return jsonify({'error': f'시트를 찾을 수 없습니다: {sheet_name}'}), 404

        ws = wb[sheet_name]
        target_row_idx = row_index + 2
        
        # 삭제 전 정보 기억 (매매일지인 경우 동기화를 위해)
        stock_name = None
        if sheet_name == '매매일지':
            stock_name = str(ws.cell(row=target_row_idx, column=2).value or "").strip()
        
        # 행 삭제
        ws.delete_rows(target_row_idx)
        
        # 매매일지 삭제 후 동기화 로직
        if sheet_name == '매매일지' and stock_name:
            # 삭제 후 시트에서 해당 종목의 마지막 거래를 다시 찾음 (날짜 순서 고려 없이 가장 아래에 있는 해당 종목 행)
            last_amount = 0
            stock_clean = stock_name.replace(" ", "")
            # 2행부터 마지막 행까지 (이미 한 행이 삭제된 상태)
            for r in range(2, ws.max_row + 1):
                cell_val = str(ws.cell(row=r, column=2).value or "").strip().replace(" ", "")
                if cell_val == stock_clean:
                    try:
                        val = ws.cell(row=r, column=6).value
                        if val is not None:
                            last_amount = float(val)
                    except:
                        pass
            
            # 포트폴리오 맵 업데이트 (마지막 찾은 금액 기준, 없으면 0)
            sync_portfolio_map(wb, stock_name, last_amount)
            print(f"🗑️ [{stock_name}] 삭제 후 이전 거래 데이터({last_amount})로 동기화 완료")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        if upload_file(file_id, output.getvalue()):
            return jsonify({'success': True, 'message': '행이 삭제되었습니다.'})
        return jsonify({'error': '파일 업로드 실패'}), 500
    except Exception as e:
        logger.error(f"Delete row error: {str(e)}")
        return jsonify({'error': f'삭제 오류: {str(e)}'}), 500


# ==================== 에러 핸들러 ====================
@app.errorhandler(401)
def unauthorized(error):
    return jsonify({
        'error': '인증 필요',
        'login_url': '/auth/login'
    }), 401


@app.errorhandler(500)
def internal_error(error):
    return jsonify({'error': '서버 오류'}), 500


if __name__ == '__main__':
    # Windows 터미널 UTF-8 출력 설정
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

    print("=" * 70)
    print("  Stock Portfolio Analysis Server - Microsoft Graph API Version")
    print("=" * 70)
    
    # 필수 설정 확인
    if not config.CLIENT_ID or not config.CLIENT_SECRET:
        print("[WARNING] CLIENT_ID와 CLIENT_SECRET이 설정되지 않았습니다.")
        print("   .env 파일을 생성하고 다음을 설정하세요:")
        print("   - TENANT_ID")
        print("   - CLIENT_ID")
        print("   - CLIENT_SECRET")
        print("   - REDIRECT_URI (선택사항, 기본값: http://localhost:5000/auth/callback)")
        print()
        print("   또는 Azure Portal에서 애플리케이션을 등록하세요.")
        print()
    
    token_data = load_token()
    if token_data:
        access_token = token_data.get('access_token')
        print("[OK] 저장된 토큰 로드됨")
    else:
        print("[INFO] 인증이 필요합니다. http://localhost:5000/auth/login 으로 이동하세요.")
    
    print(f"  Server Address: http://localhost:5000")
    print(f"  Auth Login:     http://localhost:5000/auth/login")
    print("=" * 70)
    
    try:
        app.run(debug=False, port=5000, host='0.0.0.0')
    except Exception as e:
        print(f"Error starting server: {e}")
        sys.exit(1)
