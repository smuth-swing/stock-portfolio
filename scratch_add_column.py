import os
import shutil
import openpyxl
import subprocess

EXCEL_PATH = r"C:\Users\zerod\OneDrive\주식 체크 리스트_20220328.xlsx"
BACKUP_PATH = r"C:\Users\zerod\OneDrive\주식 체크 리스트_20220328_backup.xlsx"

def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"[ERROR] Excel file not found: {EXCEL_PATH}")
        return

    # 1. 백업 파일 생성
    print(f"Creating backup: {BACKUP_PATH}")
    shutil.copy2(EXCEL_PATH, BACKUP_PATH)

    # 2. 엑셀 파일 열기
    print("Loading workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH, rich_text=True)
    
    if "탐구생활" not in wb.sheetnames:
        print("[ERROR] '탐구생활' sheet not found in the workbook.")
        return

    ws = wb["탐구생활"]
    
    # 헤더가 ["번호", "종목명", "모멘텀", "매수이유", ...] 인지 확인
    current_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"Current headers: {current_headers}")

    if "질문" in current_headers:
        print("[INFO] '질문' column already exists in '탐구생활' sheet.")
    else:
        # C열(index 3)에 열 삽입
        print("Inserting column '질문' at column index 3 (C column)...")
        ws.insert_cols(3, 1)
        
        # 헤더 설정
        ws.cell(row=1, column=3).value = "질문"
        
        # 모든 데이터 행에 빈 값 설정
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=3).value = ""
            
        print("Column inserted successfully.")

    # 3. 저장
    print("Saving workbook...")
    wb.save(EXCEL_PATH)
    wb.close()
    print("Workbook saved successfully.")

    # 4. export_to_json.py 실행하여 JSON 갱신
    print("Running export_to_json.py to update JSON files...")
    result = subprocess.run(["python", "export_to_json.py"], capture_output=True, text=True, encoding="utf-8")
    print(result.stdout)
    if result.stderr:
        print("Errors:")
        print(result.stderr)

if __name__ == "__main__":
    main()
