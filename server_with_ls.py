"""
OneDrive 엑셀 데이터 분석 서버
- OneDrive 로컬 동기화 폴더에서 엑셀 파일을 읽어 JSON API로 제공
- Flask 기반 REST API 서버
"""

import os
import sys
import json
import re
from pathlib import Path
from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS
import pandas as pd
import openpyxl
import subprocess
import threading
from openpyxl.styles import Font

# Windows 콘솔 CP949 인코딩 충돌 방지 - stdout/stderr를 UTF-8로 강제 설정
if sys.stdout.encoding != 'utf-8':
    sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)
if sys.stderr.encoding != 'utf-8':
    sys.stderr = open(sys.stderr.fileno(), mode='w', encoding='utf-8', buffering=1)


app = Flask(__name__, static_folder='.', static_url_path='')
# 모든 경로와 오리진에 대해 CORS 허용 (모바일 앱 접속용)
CORS(app, resources={r"/api/*": {"origins": "*"}})

# API 응답 gzip 압축 (대용량 JSON 전송 최적화 - 약 70~80% 크기 감소)
try:
    from flask_compress import Compress
    compress = Compress()
    app.config['COMPRESS_MIMETYPES'] = [
        'application/json',
        'text/html',
        'text/css',
        'application/javascript',
    ]
    app.config['COMPRESS_LEVEL'] = 6      # 압축 레벨 (1~9, 6이 속도/크기 균형)
    app.config['COMPRESS_MIN_SIZE'] = 500 # 500바이트 이상만 압축
    compress.init_app(app)
    print("[COMPRESS] gzip 압축 활성화됨")
except ImportError:
    print("[COMPRESS] flask-compress 미설치 — pip install flask-compress 로 설치 가능")


# 글로벌 OneDrive 경로 설정 (사용자 지정 경로 우선)
def find_target_onedrive_path():
    fixed_path = r"C:\Users\zerod\OneDrive"
    if os.path.isdir(fixed_path):
        return fixed_path
    
    # 자동 탐색 로직 (예비용)
    possible_paths = [
        os.path.expanduser("~/OneDrive"),
        os.path.expanduser("~/OneDrive - Personal"),
        os.path.expandvars(r"%USERPROFILE%\OneDrive"),
    ]
    for path in possible_paths:
        if path and os.path.isdir(path):
            return os.path.abspath(path)
    return None

ONEDRIVE_PATH = find_target_onedrive_path()

# ==================== 자동 동기화 트리거 ====================
def trigger_export():
    """데이터 변경 시 export_to_json.py를 실행하여 앱용 JSON 갱신"""
    def run():
        try:
            print("[AUTO-SYNC] JSON 변환 시작...")
            subprocess.run([sys.executable, "export_to_json.py"], check=True)
            print("[AUTO-SYNC] ✅ JSON 변환 완료!")
        except Exception as e:
            print(f"[AUTO-SYNC] ❌ 변환 실패: {e}")

    # 서버 응답을 방해하지 않도록 별도 스레드에서 실행
    threading.Thread(target=run).start()

# ==================== 유틸리티 함수 (취소선 처리) ====================
def parse_strikethrough_text(text):
    """텍스트 내의 <del>태그 또는 ~~패턴을 파싱하여 CellRichText 객체로 변환"""
    if not isinstance(text, str) or not text:
        return text
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.styles import Font
    
    text = re.sub(r'<del>(.*?)</del>', r'~~\1~~', text)
    pattern = r'~~(.*?)~~'
    import re as regex
    parts = regex.split(pattern, text)
    
    if len(parts) == 1: return text
    
    rich_text = CellRichText()
    default_font = Font(name='맑은 고딕')
    strike_font = Font(name='맑은 고딕', strike=True)
    
    for i, part in enumerate(parts):
        if not part: continue
        if i % 2 == 0:
            rich_text.add(TextBlock(default_font, part))
        else:
            rich_text.add(TextBlock(strike_font, part))
    return rich_text

def extract_rich_text(cell):
    """엑셀 셀의 실제 취소선 서식을 감지하여 마크다운(~~) 형식으로 변환"""
    from openpyxl.cell.rich_text import CellRichText
    if not hasattr(cell, 'value') or cell.value is None: return ""
    if isinstance(cell.value, CellRichText):
        result = []
        for part in cell.value:
            text = part.text if hasattr(part, 'text') else str(part)
            if hasattr(part, 'font') and part.font and part.font.strike:
                result.append(f"~~{text}~~")
            else: result.append(text)
        return "".join(result)
    if cell.font and cell.font.strike: return f"~~{cell.value}~~"
    return str(cell.value)

@app.route('/')
def index():
    """메인 페이지 서빙"""
    return send_from_directory('.', 'index.html')

# ══════════════════════════════════════════════════════════════
# LS증권 OpenAPI 연동 엔드포인트
# ══════════════════════════════════════════════════════════════

@app.route('/api/ls/config', methods=['GET', 'POST'])
def ls_config():
    """LS증권 API 설정 저장 / 조회"""
    try:
        from ls_api import load_config, save_config
    except ImportError:
        return jsonify({'error': 'ls_api 모듈을 찾을 수 없습니다.'}), 500

    if request.method == 'GET':
        cfg = load_config()
        # 비밀번호/시크릿은 마스킹하여 반환
        safe_cfg = {
            'app_key': cfg.get('app_key', ''),
            'app_secret': '****' if cfg.get('app_secret') else '',
            'account': cfg.get('account', ''),
            'account_pw': '****' if cfg.get('account_pw') else '',
            'configured': bool(cfg.get('app_key') and cfg.get('app_secret') and
                               cfg.get('account') and cfg.get('account_pw'))
        }
        return jsonify(safe_cfg)

    # POST: 설정 저장
    data = request.get_json()
    cfg = load_config()  # 기존 설정 유지 (부분 업데이트 지원)

    for key in ('app_key', 'app_secret', 'account', 'account_pw'):
        val = data.get(key, '')
        # '****'가 들어오면 기존 값 유지 (마스킹된 값 그대로 보낸 경우)
        if val and val != '****':
            cfg[key] = val

    save_config(cfg)
    return jsonify({'success': True, 'message': 'LS증권 API 설정이 저장되었습니다.'})


@app.route('/api/ls/fetch-trades', methods=['POST'])
def ls_fetch_trades():
    """
    LS증권 API로 체결 내역 조회
    Body: { from_date: "YYYYMMDD", to_date: "YYYYMMDD", stock_code: "" }
    """
    try:
        from ls_api import fetch_trade_history
    except ImportError:
        return jsonify({'error': 'ls_api 모듈을 찾을 수 없습니다.'}), 500

    data = request.get_json() or {}
    from_date = data.get('from_date', '')
    to_date = data.get('to_date', '')
    stock_code = data.get('stock_code', '')

    if not from_date or not to_date:
        return jsonify({'error': '조회 기간(from_date, to_date)을 입력하세요. 형식: YYYYMMDD'}), 400

    try:
        trades = fetch_trade_history(
            from_date=from_date,
            to_date=to_date,
            stock_code=stock_code
        )
        return jsonify({
            'success': True,
            'count': len(trades),
            'from_date': from_date,
            'to_date': to_date,
            'trades': trades
        })
    except ValueError as e:
        # 설정 미완료 등 사용자 입력 오류
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': f'LS API 조회 실패: {str(e)}'}), 500


@app.route('/api/ls/import-trades', methods=['POST'])
def ls_import_trades():
    """
    선택한 거래내역을 Excel DB(매매일지 시트)에 저장
    Body: {
        file: "파일경로",
        trades: [{ date, name, qty, price, type, investment, memo }, ...]
    }
    """
    if not ONEDRIVE_PATH:
        return jsonify({'error': 'OneDrive 경로가 설정되지 않았습니다.'}), 400

    data = request.get_json() or {}
    file_path = data.get('file', '')
    trades = data.get('trades', [])

    if not file_path:
        return jsonify({'error': 'file 경로를 지정하세요.'}), 400
    if not trades:
        return jsonify({'error': '저장할 거래 내역이 없습니다.'}), 400

    full_path = os.path.join(ONEDRIVE_PATH, file_path)
    if not os.path.isfile(full_path):
        return jsonify({'error': f'파일을 찾을 수 없습니다: {file_path}'}), 404

    try:
        wb = openpyxl.load_workbook(full_path)
        sheet_name = '매매일지'
        if sheet_name not in wb.sheetnames:
            return jsonify({'error': f'"{sheet_name}" 시트를 찾을 수 없습니다.'}), 404

        ws = wb[sheet_name]
        saved_count = 0
        errors = []

        for trade in trades:
            try:
                # 기존 매매일지 컬럼 순서: [날짜, 종목, 수량, 단가, 매매종류, 투자금]
                row_values = [
                    trade.get('date', ''),
                    trade.get('name', ''),
                    int(trade.get('qty', 0)),
                    int(trade.get('price', 0)),
                    trade.get('type', '매수'),
                    float(trade.get('investment', 0)),
                ]
                # 메모가 있으면 7번째 컬럼에 추가
                memo = trade.get('memo', '')
                if memo:
                    row_values.append(memo)

                ws.append(row_values)

                # 폰트 색상 (매도=빨강, 매수=검정)
                last_row = ws.max_row
                font_color = "FF0000" if trade.get('type') == '매도' else "000000"
                cell_font = Font(color=font_color)
                for col_idx in range(1, len(row_values) + 1):
                    ws.cell(row=last_row, column=col_idx).font = cell_font

                # 포트폴리오 맵 동기화
                stock_name = trade.get('name', '')
                investment = float(trade.get('investment', 0))
                if stock_name:
                    try:
                        sync_portfolio_map(wb, stock_name, investment)
                    except Exception:
                        pass  # 맵 동기화 실패는 저장 자체를 막지 않음

                saved_count += 1

            except Exception as row_e:
                errors.append(f"{trade.get('name', '?')}: {str(row_e)}")

        wb.save(full_path)
        wb.close()

        # 아이폰 앱용 JSON 자동 갱신
        trigger_export()

        msg = f'{saved_count}건 저장 완료'
        if errors:
            msg += f' (오류 {len(errors)}건: {"; ".join(errors[:3])})'

        return jsonify({'success': True, 'saved': saved_count, 'errors': errors, 'message': msg})

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': f'저장 오류: {str(e)}'}), 500

if __name__ == '__main__':
    print("=" * 60)
    print("  Stock Portfolio Analysis Server")
    print("=" * 60)
    if ONEDRIVE_PATH:
        print(f"  OneDrive Path: {os.path.normpath(ONEDRIVE_PATH)}")
    else:
        print("  OneDrive path not detected.")
        print("  Please set the path in the Web UI.")
    print(f"  Server Address: http://localhost:5000")
    print("=" * 60)
    
    try:
        # 배포 시에는 debug=False 권장
        app.run(debug=False, port=5000, host='0.0.0.0')
    except Exception as e:
        print(f"Error starting server: {e}")
        sys.exit(1)
