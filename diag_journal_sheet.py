# -*- coding: utf-8 -*-
"""매매일지 시트 구조 진단 (pandas + openpyxl)"""
import io, sys
import pandas as pd
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
FULL = r"C:\Users\zerod\OneDrive\주식 체크 리스트_20220328.xlsx"

with open(FULL, 'rb') as f:
    file_data = f.read()

xl = pd.ExcelFile(io.BytesIO(file_data), engine='openpyxl')
print("sheet_names:", xl.sheet_names)

target = '매매일지'
df = pd.read_excel(xl, sheet_name=target)
print("\n=== pandas head(5) (컬럼: %d개) ===" % len(df.columns))
print(df.head(5).to_string())

wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True, rich_text=True)
ws = wb[target]
print("\n=== openpyxl 원본: max_row=%d max_col=%d ===" % (ws.max_row, ws.max_col))
for r in range(1, min(8, ws.max_row) + 1):
    vals = []
    for c in range(1, min(10, ws.max_col) + 1):
        cell = ws.cell(row=r, column=c)
        v = cell.value
        vals.append(repr(v)[:40])
    print(f"Row {r}: {vals}")

print("\n=== pandas 첫 12행의 0~7열 raw 값 ===")
sub = df.iloc[:12, :8]
print(sub.to_string())
print("\ndf.iloc[0].values (repr):")
print([repr(x) for x in df.iloc[0].values[:10]])
