import openpyxl
import os

excel_path = r"C:\Users\zerod\OneDrive\주식 체크 리스트_20220328.xlsx"
wb = openpyxl.load_workbook(excel_path)

# 현금비중 시트 생성
if '현금비중' not in wb.sheetnames:
    ws = wb.create_sheet('현금비중')
    # 헤더
    ws.cell(1, 1, '월')
    ws.cell(1, 2, '투자금(백만)')
    ws.cell(1, 3, '현금(백만)')
    ws.cell(1, 4, '총자산(백만)')
    ws.cell(1, 5, '현금비중(%)')
    wb.save(excel_path)
    print(f'✅ "현금비중" 시트 생성 완료 (헤더 5개 컬럼)')
    print(f'   시트 목록: {wb.sheetnames}')
else:
    print('이미 존재함')
    ws = wb['현금비중']
    print(f'   행 수: {ws.max_row}')
    for r in range(1, min(ws.max_row+1, 5)):
        print(f'   Row {r}: {[ws.cell(r,c).value for c in range(1, 6)]}')
